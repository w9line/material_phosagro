from __future__ import annotations

import csv
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import sys
import sqlite3
import uuid
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.request import Request as URLRequest, urlopen

from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field, field_validator

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./data/app.db")
DB_PATH = DATABASE_URL.replace("sqlite:///", "", 1)
if not DATABASE_URL.startswith("postgres"):
    if not DB_PATH.startswith("/"):
        DB_PATH = str(Path(__file__).parent / DB_PATH)
    Path(DB_PATH).parent.mkdir(parents=True, exist_ok=True)


def runtime_db_config() -> tuple[str, str]:
    for module_name in ("app", "services.api.app"):
        module = sys.modules.get(module_name)
        if module is not None:
            return getattr(module, "DATABASE_URL", DATABASE_URL), getattr(module, "DB_PATH", DB_PATH)
    return DATABASE_URL, DB_PATH

MATERIALS = ("A", "B", "C")
MATERIAL_CODE = re.compile(r"^[A-Z][A-Z0-9_-]{0,15}$")
POLICIES = ("strict_fifo", "max_concentration", "hybrid")
DEFAULT_RULES = {
    "A": (28.0, 23.0, 1.0, 0.9, 0.0),
    "B": (30.0, 25.0, 1.0, 0.9, 0.0),
    "C": (35.0, 25.0, 1.0, 0.9, 0.0),
    "D": (32.0, 24.0, 1.0, 0.9, 0.0),
    "E": (28.0, 22.0, 1.0, 0.9, 0.0),
}

TOOL_REGISTRY: dict[str, dict[str, Any]] = {
    "check_batch_quality": {"title": "Качество партии", "aliases": ["проверка партии", "статус партии", "check batch quality"], "description": "Проверяет качество и статус одной партии.", "parameters": {"batch_id": {"type": "string"}}, "required": ["batch_id"], "examples": ["Проверь A-001"], "mutating": False, "category": "quality", "units": {"mass": "kg_raw"}},
    "get_batch_details": {"title": "Детали партии", "aliases": ["параметры партии", "batch details"], "description": "Возвращает фактические параметры партии.", "parameters": {"batch_id": {"type": "string"}}, "required": ["batch_id"], "examples": ["Покажи детали A-001"], "mutating": False, "category": "quality", "units": {"mass": "kg_raw"}},
    "get_oldest_batches": {"title": "Самые старые партии", "aliases": ["старые партии", "самая старая партия", "oldest batches"], "description": "Показывает партии с самой ранней датой поступления.", "parameters": {"material_type": {"type": ["string", "null"]}, "limit": {"type": "integer", "minimum": 1, "maximum": 20}}, "required": [], "examples": ["Какие самые старые партии?"], "mutating": False, "category": "inventory", "units": {"mass": "kg_raw"}},
    "build_chart": {"title": "График", "aliases": ["график", "диаграмма", "визуализация", "chart"], "description": "Строит график любой доступной измеримой метрики по партиям, качеству, браку или плану.", "parameters": {"chart_type": {"type": "string", "enum": ["inventory", "quality", "material_quality", "concentration", "metric"]}, "metric": {"type": ["string", "null"], "enum": ["raw_mass", "active_mass", "theoretical_active_mass", "recovery_loss", "concentration", "batch_count", "status_count", "status_share", "age_days", "rejection_batch_count", "rejection_raw_mass", "rejection_active_mass", "rejection_share", "required_active_mass", "plan_available_active_mass", "planned_batch_count", "covered_active_mass", "deficit_active_mass", "coverage_percent", "raw_mass_used", "loss", "safe_growth_percent", "deficit_delta", "base_deficit_active_mass", "new_deficit_active_mass", "base_rework_batch_count", "new_rework_batch_count", None]}, "group_by": {"type": ["string", "null"], "enum": ["material", "status", "batch", "arrival_date", "policy", None]}, "material_type": {"type": ["string", "null"]}, "requirements": {"type": ["object", "null"]}, "policy": {"type": ["string", "null"], "enum": [*POLICIES, None]}, "changes_percent": {"type": ["object", "null"]}, "policies": {"type": ["array", "null"], "items": {"type": "string", "enum": list(POLICIES)}}}, "required": [], "examples": ["Построй график остатков"], "mutating": False, "category": "reports", "units": {"raw_mass": "kg_raw", "active_mass": "kg_active", "concentration": "percent", "age": "days", "share": "percent"}},
    "classify_batches": {"title": "Классификация", "aliases": ["классификац", "классификация партий", "classify batches"], "description": "Классифицирует партии по действующим порогам качества.", "parameters": {"material_type": {"type": ["string", "null"]}, "only_unclassified": {"type": "boolean"}}, "required": [], "examples": ["Классифицируй партии A"], "mutating": False, "category": "quality", "units": {}, "filters": ["material_type"]},
    "get_inventory_summary": {"title": "Остатки", "aliases": ["остатков", "склад", "запасы", "сырьё", "inventory"], "description": "Показывает остатки сырой массы и доступного активного вещества.", "parameters": {"material_type": {"type": ["string", "null"]}, "group_by": {"type": "string", "enum": ["material_and_status"]}}, "required": [], "examples": ["Покажи остатки по A"], "mutating": False, "category": "inventory", "units": {"raw_mass": "kg_raw", "active_mass": "kg_active"}, "filters": ["material_type"]},
    "build_weekly_plan": {"title": "Недельный план", "aliases": ["план производства", "weekly plan", "планирование"], "description": "Строит безопасный preview-план без списания остатков.", "parameters": {"requirements": {"type": "object"}, "policy": {"type": "string", "enum": list(POLICIES)}, "allow_rework": {"type": "boolean"}}, "required": ["requirements", "policy"], "examples": ["Построй план A 3000 B 2500 C 1800, hybrid"], "mutating": False, "category": "planning", "units": {"requirements": "kg_active"}},
    "check_material_deficit": {"title": "Дефицит", "aliases": ["не хватит", "потребность", "material deficit"], "description": "Сравнивает потребность с доступным активным веществом.", "parameters": {"requirements": {"type": "object"}, "include_rework": {"type": "boolean"}}, "required": ["requirements"], "examples": ["Проверь дефицит A 3000"], "mutating": False, "category": "planning", "units": {"requirements": "kg_active", "deficit": "kg_active"}},
    "compare_allocation_policies": {"title": "Стратегии", "aliases": ["политики распределения", "сравнение стратегий", "compare policies", "allocation policies", "fifo"], "description": "Сравнивает FIFO, max concentration и hybrid без изменения данных.", "parameters": {"requirements": {"type": "object"}, "policies": {"type": "array", "items": {"type": "string", "enum": list(POLICIES)}}}, "required": ["requirements", "policies"], "examples": ["Сравни FIFO и hybrid для A 3000"], "mutating": False, "category": "planning", "units": {"requirements": "kg_active"}},
    "generate_rejection_report": {"title": "Отчёт по браку", "aliases": ["отчёт по отклонениям", "rejection report", "брак"], "description": "Формирует read-only отчёт по REWORK и REJECTED.", "parameters": {"material_type": {"type": ["string", "null"]}, "include_rework": {"type": "boolean"}, "include_rejected": {"type": "boolean"}}, "required": [], "examples": ["Сделай отчёт по браку A"], "mutating": False, "category": "reports", "units": {"mass": "kg_raw"}, "filters": ["material_type"]},
    "simulate_requirement_change": {"title": "Сценарий потребности", "aliases": ["изменение спроса", "simulate requirement", "сценарий"], "description": "Сравнивает базовый и изменённый спрос без изменения склада.", "parameters": {"base_requirements": {"type": "object"}, "changes_percent": {"type": "object"}, "policy": {"type": "string", "enum": list(POLICIES)}}, "required": ["changes_percent", "policy"], "examples": ["Что будет, если B вырастет на 20%"], "mutating": False, "category": "planning", "units": {"requirements": "kg_active", "changes": "percent"}},
}


class CompatRow(dict):
    def __getitem__(self, key: Any) -> Any:
        return list(self.values())[key] if isinstance(key, int) else super().__getitem__(key)


class PostgresCursor:
    def __init__(self, cursor: Any):
        self.cursor = cursor

    @property
    def rowcount(self) -> int:
        return self.cursor.rowcount

    def _row(self, row: Any) -> Any:
        if row is None or not self.cursor.description:
            return row
        return CompatRow(zip([column.name for column in self.cursor.description], row))

    def fetchone(self) -> Any:
        return self._row(self.cursor.fetchone())

    def fetchall(self) -> list[Any]:
        return [self._row(row) for row in self.cursor.fetchall()]

    def __iter__(self):
        for row in self.cursor:
            yield self._row(row)


class PostgresConnection:
    def __init__(self, url: str):
        import psycopg
        self.con = psycopg.connect(url)

    def execute(self, sql: str, params: tuple[Any, ...] = ()) -> Any:
        cursor = self.con.cursor(); cursor.execute(sql.replace("?", "%s"), params); return PostgresCursor(cursor)

    def executemany(self, sql: str, params: Any) -> Any:
        cursor = self.con.cursor(); cursor.executemany(sql.replace("?", "%s"), params); return PostgresCursor(cursor)

    def executescript(self, script: str) -> None:
        for statement in script.split(";"):
            if statement.strip():
                self.execute(statement)

    def commit(self) -> None:
        self.con.commit()

    def rollback(self) -> None:
        self.con.rollback()

    def close(self) -> None:
        self.con.close()


def db() -> Any:
    database_url, db_path = runtime_db_config()
    if database_url.startswith("postgres"):
        return PostgresConnection(database_url)
    con = sqlite3.connect(db_path, check_same_thread=False)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys=ON")
    return con


def material_codes() -> tuple[str, ...]:
    con = db()
    rows = con.execute("SELECT material_type FROM quality_rules").fetchall()
    batches = con.execute("SELECT DISTINCT material_type FROM batches").fetchall()
    requirements = con.execute("SELECT material_type FROM requirements").fetchall()
    con.close()
    codes = {row[0] for row in rows + batches + requirements if row[0]}
    return tuple(sorted(codes or set(MATERIALS)))


def required_material_codes() -> tuple[str, ...]:
    con = db(); rows = con.execute("SELECT material_type FROM requirements WHERE required_active_mass_kg > 0").fetchall(); con.close()
    return tuple(row[0] for row in rows) or MATERIALS


def ensure_materials(con: Any, codes: list[str] | tuple[str, ...] | set[str]) -> None:
    known = material_codes_from_connection(con)
    for code in codes:
        code = code.upper()
        if not MATERIAL_CODE.fullmatch(code):
            raise ValueError(f"invalid material code: {code}")
        if code not in known:
            values = DEFAULT_RULES.get(code, DEFAULT_RULES["A"])
            con.execute("INSERT INTO quality_rules VALUES (?,?,?,?,?,?) ON CONFLICT(material_type) DO NOTHING", (code, *values))
            con.execute("INSERT INTO requirements VALUES (?,?) ON CONFLICT(material_type) DO NOTHING", (code, 0.0))
            known.add(code)


def material_codes_from_connection(con: Any) -> set[str]:
    return {row[0] for row in con.execute("SELECT material_type FROM quality_rules").fetchall()}


def bump_data_version(con: Any) -> int:
    row = con.execute("SELECT value FROM settings WHERE key='data_version'").fetchone()
    version = int(row[0] if row else 0) + 1
    con.execute("INSERT INTO settings(key, value) VALUES ('data_version', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (str(version),))
    return version


def snapshot_meta(units: dict[str, str], parameters: dict[str, Any] | None = None) -> dict[str, Any]:
    con = db(); row = con.execute("SELECT value FROM settings WHERE key='data_version'").fetchone(); con.close()
    return {"data_version": int(row[0]) if row else 1, "calculated_at": datetime.utcnow().isoformat() + "Z", "parameters": parameters or {}, "units": units}


def init_db() -> None:
    con = db()
    con.executescript("""
    CREATE TABLE IF NOT EXISTS batches (
      batch_id TEXT PRIMARY KEY, material_type TEXT NOT NULL, raw_mass_kg REAL NOT NULL,
      concentration_percent REAL NOT NULL, arrival_date TEXT NOT NULL, supplier TEXT,
      warehouse TEXT, certificate_number TEXT, notes TEXT, remaining_raw_mass_kg REAL NOT NULL,
      source TEXT NOT NULL, status TEXT, created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS quality_rules (
      material_type TEXT PRIMARY KEY, good_threshold_percent REAL NOT NULL,
      rework_threshold_percent REAL NOT NULL, good_recovery_factor REAL NOT NULL,
      rework_recovery_factor REAL NOT NULL, reject_recovery_factor REAL NOT NULL
    );
    CREATE TABLE IF NOT EXISTS requirements (material_type TEXT PRIMARY KEY, required_active_mass_kg REAL NOT NULL);
    CREATE TABLE IF NOT EXISTS plans (plan_id TEXT PRIMARY KEY, status TEXT NOT NULL, payload TEXT NOT NULL, created_at TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS users (
      user_id TEXT PRIMARY KEY, username TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
      is_admin INTEGER NOT NULL DEFAULT 0, is_blocked INTEGER NOT NULL DEFAULT 0, created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS sessions (
      token_hash TEXT PRIMARY KEY, user_id TEXT NOT NULL, expires_at TEXT NOT NULL,
      FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS plan_owners (plan_id TEXT PRIMARY KEY, user_id TEXT NOT NULL, status TEXT NOT NULL DEFAULT 'preview', created_at TEXT NOT NULL, preview_data_version INTEGER NOT NULL DEFAULT 1, FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE);
    CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS chats (
      chat_id TEXT PRIMARY KEY, user_id TEXT NOT NULL, title TEXT NOT NULL,
      created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
      FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS chat_messages (
      message_id TEXT PRIMARY KEY, chat_id TEXT NOT NULL, role TEXT NOT NULL,
      content TEXT NOT NULL, tool_calls TEXT NOT NULL DEFAULT '[]', created_at TEXT NOT NULL,
      FOREIGN KEY(chat_id) REFERENCES chats(chat_id) ON DELETE CASCADE
    );
    INSERT INTO settings(key, value) VALUES ('registration_open', '1') ON CONFLICT(key) DO NOTHING;
    INSERT INTO settings(key, value) VALUES ('data_version', '1') ON CONFLICT(key) DO NOTHING;
    """)
    if DATABASE_URL.startswith("postgres"):
        column_exists = con.execute("SELECT 1 FROM information_schema.columns WHERE table_name='plan_owners' AND column_name='preview_data_version'").fetchone()
    else:
        column_exists = next((row for row in con.execute("PRAGMA table_info(plan_owners)").fetchall() if row[1] == "preview_data_version"), None)
    if not column_exists:
        con.execute("ALTER TABLE plan_owners ADD COLUMN preview_data_version INTEGER NOT NULL DEFAULT 1")
    for material, values in DEFAULT_RULES.items():
        con.execute("INSERT INTO quality_rules VALUES (?,?,?,?,?,?) ON CONFLICT(material_type) DO NOTHING", (material, *values))
        con.execute("INSERT INTO requirements VALUES (?,?) ON CONFLICT(material_type) DO NOTHING", (material, 3000.0))
    if con.execute("SELECT COUNT(*) FROM batches").fetchone()[0] == 0:
        demo = [
            ("A-001", "A", 5000, 24.5, "2026-08-01", "Поставщик-1", "Основной", None, "Демо", 5000, "demo", None, datetime.utcnow().isoformat()),
            ("A-002", "A", 4200, 30.1, "2026-08-03", "Поставщик-2", "Основной", None, "Демо", 4200, "demo", None, datetime.utcnow().isoformat()),
            ("B-001", "B", 7000, 38.2, "2026-08-02", "Поставщик-1", "Зона B", None, "Демо", 7000, "demo", None, datetime.utcnow().isoformat()),
            ("C-001", "C", 6000, 32.0, "2026-08-04", "Поставщик-3", "Зона C", None, "Демо", 6000, "demo", None, datetime.utcnow().isoformat()),
        ]
        con.executemany("INSERT INTO batches VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", demo)
    admin_username = os.getenv("ADMIN_USERNAME", "admin").strip()
    admin_password = os.getenv("ADMIN_PASSWORD", "")
    if admin_password and not con.execute("SELECT 1 FROM users WHERE username=?", (admin_username,)).fetchone():
        con.execute("INSERT INTO users VALUES (?,?,?,?,?,?)", (str(uuid.uuid4()), admin_username, hash_password(admin_password), 1, 0, datetime.utcnow().isoformat()))
    con.commit()
    con.close()


class BatchIn(BaseModel):
    batch_id: str
    material_type: str
    raw_mass_kg: float
    concentration_percent: float
    arrival_date: date
    supplier: str | None = None
    warehouse: str | None = None
    certificate_number: str | None = None
    notes: str | None = None
    remaining_raw_mass_kg: float | None = None
    source: str = "manual"

    @field_validator("material_type")
    @classmethod
    def material_ok(cls, value: str) -> str:
        value = value.upper()
        if not MATERIAL_CODE.fullmatch(value):
            raise ValueError("material_type must start with A-Z and contain up to 16 safe characters")
        return value

    @field_validator("batch_id")
    @classmethod
    def id_ok(cls, value: str) -> str:
        if not value.strip():
            raise ValueError("batch_id is required")
        return value.strip()

    @field_validator("raw_mass_kg")
    @classmethod
    def mass_ok(cls, value: float) -> float:
        if value <= 0:
            raise ValueError("raw_mass_kg must be positive")
        return value

    @field_validator("concentration_percent")
    @classmethod
    def concentration_ok(cls, value: float) -> float:
        if not 0 <= value <= 100:
            raise ValueError("concentration_percent must be 0..100")
        return value


class RequirementIn(BaseModel):
    requirements: dict[str, float]
    policy: str = "hybrid"
    allow_rework: bool = True

    @field_validator("requirements")
    @classmethod
    def requirements_ok(cls, value: dict[str, float]) -> dict[str, float]:
        if any(not MATERIAL_CODE.fullmatch(material.upper()) or amount < 0 for material, amount in value.items()):
            raise ValueError("requirements must contain non-negative values for valid material codes")
        return {material.upper(): float(amount) for material, amount in value.items()}

    @field_validator("policy")
    @classmethod
    def policy_ok(cls, value: str) -> str:
        if value not in POLICIES: raise ValueError(f"policy must be one of {POLICIES}")
        return value


class RuleIn(BaseModel):
    good_threshold_percent: float
    rework_threshold_percent: float
    good_recovery_factor: float = 1.0
    rework_recovery_factor: float = 0.9
    reject_recovery_factor: float = 0.0


class ChatIn(BaseModel):
    message: str = Field(min_length=1, max_length=4000)
    chat_id: str | None = None
    history: list[dict[str, str]] = Field(default_factory=list, max_length=20)


class AuthIn(BaseModel):
    username: str = Field(min_length=3, max_length=32)
    password: str = Field(min_length=8, max_length=128)


class UserPatch(BaseModel):
    is_blocked: bool | None = None
    is_admin: bool | None = None


class RegistrationSetting(BaseModel):
    registration_open: bool


def hash_password(password: str, salt: bytes | None = None) -> str:
    salt = salt or secrets.token_bytes(16)
    rounds = 240_000
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, rounds)
    return f"pbkdf2${rounds}${salt.hex()}${digest.hex()}"


def verify_password(password: str, encoded: str) -> bool:
    try:
        scheme, rounds, salt_hex, digest_hex = encoded.split("$", 3)
        if scheme != "pbkdf2":
            return False
        digest = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt_hex), int(rounds))
        return hmac.compare_digest(digest.hex(), digest_hex)
    except (ValueError, TypeError):
        return False


def public_user(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    return {"user_id": row["user_id"], "username": row["username"], "is_admin": bool(row["is_admin"]), "is_blocked": bool(row["is_blocked"]), "created_at": row["created_at"]}


def session_token(con: sqlite3.Connection, user_id: str) -> str:
    raw = secrets.token_urlsafe(32)
    con.execute("INSERT INTO sessions VALUES (?,?,?)", (hashlib.sha256(raw.encode()).hexdigest(), user_id, datetime.fromtimestamp(datetime.now().timestamp() + 60 * 60 * 24 * 30).isoformat()))
    return raw


def current_user(request: Request, admin: bool = False) -> sqlite3.Row:
    header = request.headers.get("authorization", "")
    token = header[7:] if header.lower().startswith("bearer ") else request.headers.get("x-session-token", "")
    if not token:
        raise HTTPException(401, "authentication required")
    con = db(); row = con.execute("SELECT u.* FROM sessions s JOIN users u ON u.user_id=s.user_id WHERE s.token_hash=? AND s.expires_at>?", (hashlib.sha256(token.encode()).hexdigest(), datetime.utcnow().isoformat())).fetchone()
    if not row:
        con.close(); raise HTTPException(401, "invalid or expired session")
    if row["is_blocked"]:
        con.close(); raise HTTPException(403, "account is blocked")
    if admin and not row["is_admin"]:
        con.close(); raise HTTPException(403, "admin access required")
    con.close()
    return row


def chat_for_user(con: sqlite3.Connection, chat_id: str, user_id: str) -> sqlite3.Row:
    row = con.execute("SELECT * FROM chats WHERE chat_id=? AND user_id=?", (chat_id, user_id)).fetchone()
    if not row:
        raise HTTPException(404, "chat not found")
    return row


def save_message(con: sqlite3.Connection, chat_id: str, role: str, content: str, tool_calls: list[dict[str, Any]] | None = None) -> None:
    now = datetime.utcnow().isoformat()
    con.execute("INSERT INTO chat_messages VALUES (?,?,?,?,?,?)", (str(uuid.uuid4()), chat_id, role, content, json.dumps(tool_calls or [], ensure_ascii=False), now))
    con.execute("UPDATE chats SET updated_at=? WHERE chat_id=?", (now, chat_id))


def trace_with_result(trace: list[dict[str, Any]], data: Any) -> list[dict[str, Any]]:
    return [{**item, **({"result": data.get("chart", data) if isinstance(data, dict) else data} if item.get("tool") == "build_chart" else {})} for item in trace]


def rules() -> dict[str, dict[str, float]]:
    con = db(); rows = con.execute("SELECT * FROM quality_rules").fetchall(); con.close()
    return {r["material_type"]: dict(r) for r in rows}


def classify(row: dict[str, Any]) -> dict[str, Any]:
    rule = rules()[row["material_type"]]
    c = float(row["concentration_percent"])
    if c >= rule["good_threshold_percent"]:
        status, recovery = "GOOD", rule["good_recovery_factor"]
    elif c >= rule["rework_threshold_percent"]:
        status, recovery = "REWORK", rule["rework_recovery_factor"]
    else:
        status, recovery = "REJECTED", rule["reject_recovery_factor"]
    return {"status": status, "recovery_factor": recovery, "reason": {"GOOD": "Концентрация соответствует норме", "REWORK": "Концентрация ниже нормы, но партия пригодна для доработки", "REJECTED": "Концентрация ниже порога доработки"}[status]}


def batch_dict(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    result = dict(row)
    result["quality"] = classify(result)
    result["theoretical_active_mass_kg"] = round(result["raw_mass_kg"] * result["concentration_percent"] / 100, 3)
    result["available_active_mass_kg"] = round(result["remaining_raw_mass_kg"] * result["concentration_percent"] / 100 * result["quality"]["recovery_factor"], 3)
    result["meta"] = snapshot_meta({"raw_mass": "kg_raw", "active_mass": "kg_active"}, {"batch_id": result["batch_id"]})
    return result


def validate_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    valid, errors, seen = [], [], set()
    con = db(); existing = {r[0] for r in con.execute("SELECT batch_id FROM batches")}; con.close()
    for number, raw in enumerate(rows, 2):
        try:
            item = BatchIn.model_validate(raw)
            if item.batch_id in seen or item.batch_id in existing:
                raise ValueError("batch_id must be unique")
            remaining = item.raw_mass_kg if item.remaining_raw_mass_kg is None else item.remaining_raw_mass_kg
            if remaining < 0 or remaining > item.raw_mass_kg:
                raise ValueError("remaining_raw_mass_kg must be between 0 and raw_mass_kg")
            data = item.model_dump(mode="json"); data["remaining_raw_mass_kg"] = remaining
            valid.append(data); seen.add(item.batch_id)
        except Exception as exc:
            errors.append({"row": number, "field": "row", "code": "INVALID_VALUE", "message": str(exc)})
    return valid, errors


def save_batches(rows: list[dict[str, Any]]) -> int:
    con = db()
    ensure_materials(con, {item["material_type"] for item in rows})
    for item in rows:
        con.execute("INSERT INTO batches VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", (
            item["batch_id"], item["material_type"], item["raw_mass_kg"], item["concentration_percent"],
            item["arrival_date"], item.get("supplier"), item.get("warehouse"), item.get("certificate_number"),
            item.get("notes"), item["remaining_raw_mass_kg"], item.get("source", "upload"), None, datetime.utcnow().isoformat()))
    bump_data_version(con); con.commit(); con.close(); return len(rows)


def parse_file(content: bytes, filename: str) -> list[dict[str, Any]]:
    if filename.lower().endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            name = next((n for n in archive.namelist() if n.lower().endswith("batches.csv")), None)
            if not name: raise ValueError("ZIP must contain batches.csv")
            content, filename = archive.read(name), name
    if filename.lower().endswith(".xlsx"):
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = book[book.sheetnames[0]]
        values = list(sheet.values); headers = [str(x) for x in values[0]]
        return [dict(zip(headers, row)) for row in values[1:] if any(x is not None for x in row)]
    text = content.decode("utf-8-sig")
    sample = text[:4096]
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    return list(csv.DictReader(io.StringIO(text), dialect=dialect))


def sort_batches(rows: list[dict[str, Any]], policy: str, allow_rework: bool) -> list[dict[str, Any]]:
    allowed = [x for x in rows if x["quality"]["status"] == "GOOD" or (allow_rework and x["quality"]["status"] == "REWORK")]
    if policy == "strict_fifo":
        return sorted(allowed, key=lambda x: (x["arrival_date"], x["created_at"], x["batch_id"]))
    if policy == "max_concentration":
        return sorted(allowed, key=lambda x: (x["quality"]["status"] != "GOOD", -x["concentration_percent"], x["arrival_date"], x["batch_id"]))
    good = sorted([x for x in allowed if x["quality"]["status"] == "GOOD"], key=lambda x: (x["arrival_date"], x["created_at"], x["batch_id"]))
    rework = sorted([x for x in allowed if x["quality"]["status"] == "REWORK"], key=lambda x: (-x["concentration_percent"], x["arrival_date"], x["batch_id"]))
    return good + rework


def build_plan(requirements: dict[str, float], policy: str = "hybrid", allow_rework: bool = True) -> dict[str, Any]:
    if policy not in POLICIES: raise ValueError(f"policy must be one of {POLICIES}")
    con = db(); all_rows = [batch_dict(r) for r in con.execute("SELECT * FROM batches WHERE remaining_raw_mass_kg > 0").fetchall()]; con.close()
    materials = tuple(sorted(set(material_codes()) | set(requirements)))
    by_material = {m: [r for r in all_rows if r["material_type"] == m] for m in materials}
    result = {"policy": policy, "requirements": requirements, "materials": {}, "warnings": [], "meta": snapshot_meta({"requirements": "kg_active", "raw_mass": "kg_raw", "active_mass": "kg_active"}, {"policy": policy, "allow_rework": allow_rework, "requirements": requirements})}
    for material in materials:
        required = max(0.0, float(requirements.get(material, 0)))
        need = required; items = []; available = 0.0
        for selection_rank, row in enumerate(sort_batches(by_material[material], policy, allow_rework), 1):
            factor = row["concentration_percent"] / 100 * row["quality"]["recovery_factor"]
            capacity = row["remaining_raw_mass_kg"] * factor; available += capacity
            if need <= 1e-9: continue
            active = min(need, capacity); raw = active / factor if factor else 0
            if policy == "strict_fifo":
                selection_reason = f"Самая ранняя доступная {row['quality']['status']}-партия по FIFO"
            elif policy == "max_concentration":
                selection_reason = f"Максимальная концентрация среди доступных {row['quality']['status']}-партий"
            elif row["quality"]["status"] == "GOOD":
                selection_reason = "GOOD-партия выбрана по FIFO до использования REWORK"
            else:
                selection_reason = "REWORK-партия выбрана после GOOD по убыванию концентрации"
            items.append({"batch_id": row["batch_id"], "status": row["quality"]["status"], "selection_rank": selection_rank, "selection_reason": selection_reason, "arrival_date": row["arrival_date"], "concentration_percent": row["concentration_percent"], "raw_mass_used_kg": round(raw, 3), "active_mass_kg": round(active, 3), "active_mass_received_kg": round(active, 3), "loss_kg": round(raw * row["concentration_percent"] / 100 * (1 - row["quality"]["recovery_factor"]), 3)})
            need -= active
        covered = required - max(0, need)
        result["materials"][material] = {"required_active_mass_kg": round(required, 3), "available_active_mass_kg": round(available, 3), "covered_active_mass_kg": round(covered, 3), "deficit_active_mass_kg": round(max(0, need), 3), "coverage_percent": round(covered / required * 100, 2) if required else 100.0, "items": items, "raw_mass_used_kg": round(sum(i["raw_mass_used_kg"] for i in items), 3), "loss_kg": round(sum(i["loss_kg"] for i in items), 3), "selection_explanation": f"Стратегия {policy}; отобрано партий: {len(items)}"}
        if need > 1e-9: result["warnings"].append(f"Дефицит материала {material}: {need:.3f} кг активного вещества")
    return result


def requirements_default() -> dict[str, float]:
    con = db(); rows = con.execute("SELECT * FROM requirements").fetchall(); con.close(); return {r[0]: r[1] for r in rows}


CHART_METRIC_LABELS = {
    "raw_mass": ("Сырьё, кг", "kg"), "active_mass": ("Доступное активное вещество, кг", "kg_active"),
    "theoretical_active_mass": ("Теоретическое активное вещество, кг", "kg_active"), "recovery_loss": ("Потери восстановления, кг", "kg_active"),
    "concentration": ("Концентрация, %", "%"), "batch_count": ("Количество партий", "batches"),
    "status_count": ("Количество партий по качеству", "batches"), "status_share": ("Доля партий по качеству, %", "%"),
    "age_days": ("Возраст партий, дней", "days"), "rejection_batch_count": ("Проблемные партии", "batches"),
    "rejection_raw_mass": ("Брак и доработка, кг сырья", "kg_raw"), "rejection_active_mass": ("Брак и доработка, кг активного вещества", "kg_active"),
    "rejection_share": ("Доля брака и доработки, %", "%"), "required_active_mass": ("Требуемое активное вещество, кг", "kg_active"),
    "plan_available_active_mass": ("Доступно для плана, кг активного вещества", "kg_active"), "planned_batch_count": ("Выбранные партии", "batches"),
    "covered_active_mass": ("Покрыто активного вещества, кг", "kg_active"),
    "deficit_active_mass": ("Дефицит активного вещества, кг", "kg_active"), "coverage_percent": ("Покрытие, %", "%"),
    "raw_mass_used": ("Использовано сырья, кг", "kg_raw"), "loss": ("Потери плана, кг", "kg_raw"),
    "safe_growth_percent": ("Безопасный рост потребности, %", "%"), "deficit_delta": ("Изменение дефицита, кг", "kg_active"),
    "base_deficit_active_mass": ("Базовый дефицит, кг", "kg_active"), "new_deficit_active_mass": ("Новый дефицит, кг", "kg_active"),
    "base_rework_batch_count": ("Базовые REWORK-партии", "batches"), "new_rework_batch_count": ("Новые REWORK-партии", "batches"),
}


def _chart_row_value(row: dict[str, Any], metric: str) -> float:
    raw = float(row.get("remaining_raw_mass_kg", 0) or 0)
    concentration = float(row.get("concentration_percent", 0) or 0)
    active = float(row.get("available_active_mass_kg", 0) or 0)
    return {"raw_mass": raw, "active_mass": active, "theoretical_active_mass": raw * concentration / 100, "recovery_loss": raw * concentration / 100 - active, "concentration": concentration, "age_days": max(0, (date.today() - date.fromisoformat(str(row["arrival_date"]))).days)}.get(metric, 0.0)


def _chart_group_key(row: dict[str, Any], group_by: str | None) -> str:
    if group_by in (None, "material"): return str(row["material_type"])
    if group_by == "status": return str(row["quality"]["status"])
    if group_by == "batch": return str(row["batch_id"])
    return str(row["arrival_date"])


def _chart_from_plan(plan: dict[str, Any], metric: str, group_by: str | None, title: str, unit: str) -> tuple[list[str], list[dict[str, Any]]]:
    value_key = {"required_active_mass": "required_active_mass_kg", "plan_available_active_mass": "available_active_mass_kg", "covered_active_mass": "covered_active_mass_kg", "deficit_active_mass": "deficit_active_mass_kg", "coverage_percent": "coverage_percent", "raw_mass_used": "raw_mass_used_kg", "loss": "loss_kg"}.get(metric, metric)
    if group_by == "batch":
        items = [(item["batch_id"], item) for value in plan["materials"].values() for item in value.get("items", [])]
        item_key = "active_mass_kg" if metric == "covered_active_mass" else "raw_mass_used_kg" if metric == "raw_mass_used" else "loss_kg"
        return [label for label, _ in items], [{"name": title, "values": [round(float(item.get(item_key, 0) or 0), 3) for _, item in items]}]
    if group_by in ("status", "arrival_date") and metric in {"covered_active_mass", "raw_mass_used", "loss"}:
        items = [(item["status"] if group_by == "status" else str(item["arrival_date"]), item) for value in plan["materials"].values() for item in value.get("items", [])]
        labels = sorted({label for label, _ in items})
        item_key = "active_mass_kg" if metric == "covered_active_mass" else "raw_mass_used_kg" if metric == "raw_mass_used" else "loss_kg"
        return labels, [{"name": title, "values": [round(sum(float(item.get(item_key, 0) or 0) for label, item in items if label == current), 3) for current in labels]}]
    labels = sorted(plan["materials"])
    if metric == "planned_batch_count": return labels, [{"name": title, "values": [len(plan["materials"][label].get("items", [])) for label in labels]}]
    return labels, [{"name": title, "values": [round(float(plan["materials"][label].get(value_key, 0) or 0), 3) for label in labels]}]


def _chart_payload(rows: list[dict[str, Any]], metric: str, group_by: str | None, material: str | None) -> tuple[list[str], list[dict[str, Any]], str, str]:
    labels = sorted({_chart_group_key(row, group_by) for row in rows})
    title, unit = CHART_METRIC_LABELS[metric]
    if metric in ("status_count", "status_share"):
        statuses = ("GOOD", "REWORK", "REJECTED")
        if group_by == "status":
            values = [sum(row["quality"]["status"] == status for row in rows) for status in labels]
            if metric == "status_share": values = [round(value / max(1, len(rows)) * 100, 2) for value in values]
            return labels, [{"name": "Партий" if metric == "status_count" else "Доля, %", "values": values}], title, unit
        if group_by not in (None, "material"):
            values = [sum(_chart_group_key(row, group_by) == label for row in rows) for label in labels]
            if metric == "status_share": values = [round(value / max(1, len(rows)) * 100, 2) for value in values]
            return labels, [{"name": "Партий" if metric == "status_count" else "Доля, %", "values": values}], title, unit
        series = []
        for status in statuses:
            values = [sum(row["material_type"] == label and row["quality"]["status"] == status for row in rows) for label in labels]
            if metric == "status_share": values = [round(value / max(1, sum(row["material_type"] == label for row in rows)) * 100, 2) for label, value in zip(labels, values)]
            series.append({"name": status, "values": values})
        return labels, series, title, unit
    if metric in ("rejection_batch_count", "rejection_raw_mass", "rejection_active_mass", "rejection_share"):
        selected = [row for row in rows if row["quality"]["status"] in ("REWORK", "REJECTED")]
        values = []
        for label in labels:
            selected_group = [row for row in selected if _chart_group_key(row, group_by) == label]
            all_group = [row for row in rows if _chart_group_key(row, group_by) == label]
            if metric == "rejection_batch_count": value = len(selected_group)
            elif metric == "rejection_raw_mass": value = sum(row.get("remaining_raw_mass_kg", 0) for row in selected_group)
            elif metric == "rejection_active_mass": value = sum(row.get("available_active_mass_kg", 0) for row in selected_group)
            else: value = len(selected_group) / max(1, len(all_group)) * 100
            values.append(value)
        return labels, [{"name": title, "values": [round(value, 3) for value in values]}], title, unit
    values = [sum(_chart_row_value(row, metric) for row in rows if _chart_group_key(row, group_by) == label) for label in labels]
    if metric == "concentration": values = [round(sum(_chart_row_value(row, metric) for row in rows if _chart_group_key(row, group_by) == label) / max(1, sum(_chart_group_key(row, group_by) == label for row in rows)), 2) for label in labels]
    if metric == "batch_count": values = [sum(_chart_group_key(row, group_by) == label for row in rows) for label in labels]
    return labels, [{"name": title, "values": [round(value, 3) for value in values]}], title, unit


def tool(name: str, args: dict[str, Any]) -> Any:
    con = db()
    if name == "check_batch_quality":
        row = con.execute("SELECT * FROM batches WHERE LOWER(batch_id)=LOWER(?)", (args["batch_id"],)).fetchone(); con.close()
        if not row: raise ValueError("batch not found")
        return batch_dict(row)
    if name == "get_batch_details":
        row = con.execute("SELECT * FROM batches WHERE LOWER(batch_id)=LOWER(?)", (args["batch_id"],)).fetchone(); con.close()
        if not row: raise ValueError("batch not found")
        return batch_dict(row)
    if name == "get_oldest_batches":
        material = args.get("material_type"); limit = min(20, max(1, int(args.get("limit", 5))))
        rows = con.execute("SELECT * FROM batches" + (" WHERE material_type=?" if material else "") + " ORDER BY arrival_date, created_at, batch_id LIMIT ?", (material, limit) if material else (limit,)).fetchall(); con.close()
        return {"batches": [batch_dict(row) for row in rows], "limit": limit, "material_type": material, "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], {"material_type": material, "limit": limit})}
    if name == "build_chart":
        material = args.get("material_type"); chart_type = args.get("chart_type", "inventory"); metric = args.get("metric")
        legacy_metrics = {"inventory": "raw_mass", "quality": "status_count", "material_quality": "status_count", "concentration": "concentration"}
        metric = metric or legacy_metrics.get(chart_type, "raw_mass")
        group_by = args.get("group_by") or ("status" if chart_type == "quality" else "material")
        plan_metrics = {"required_active_mass", "plan_available_active_mass", "planned_batch_count", "covered_active_mass", "deficit_active_mass", "coverage_percent", "raw_mass_used", "loss", "safe_growth_percent", "deficit_delta", "base_deficit_active_mass", "new_deficit_active_mass", "base_rework_batch_count", "new_rework_batch_count"}
        if metric in plan_metrics:
            con.close()
            title, unit = CHART_METRIC_LABELS[metric]
            requirements = args.get("requirements") or requirements_default()
            policy = args.get("policy") or "hybrid"
            if group_by == "policy":
                policies = args.get("policies") or list(POLICIES)
                labels = list(policies)
                if metric in {"safe_growth_percent", "deficit_delta", "base_deficit_active_mass", "new_deficit_active_mass", "base_rework_batch_count", "new_rework_batch_count"}:
                    changes = args.get("changes_percent") or {}
                    if not changes: raise ValueError("Для графика сценария нужны changes_percent")
                    scenarios = {name: tool("simulate_requirement_change", {"base_requirements": requirements, "changes_percent": changes, "policy": name}) for name in policies}
                    key = {"safe_growth_percent": "safe_growth_percent", "deficit_delta": "deficit_delta_active_mass_kg", "base_deficit_active_mass": "base_deficit_active_mass_kg", "new_deficit_active_mass": "new_deficit_active_mass_kg", "base_rework_batch_count": "base_rework_batches", "new_rework_batch_count": "new_rework_batches"}[metric]
                    values = []
                    for name in labels:
                        numbers = [float(item["comparison"][material].get(key, 0) or 0) for item in scenarios[name]["comparison"]]
                        values.append(round(sum(numbers) / max(1, len(numbers)) if metric == "safe_growth_percent" else sum(numbers), 3))
                    series = [{"name": title, "values": values}]
                else:
                    comparisons = {name: build_plan(requirements, name, True) for name in policies}
                    key = {"required_active_mass": "required_active_mass_kg", "plan_available_active_mass": "available_active_mass_kg", "covered_active_mass": "covered_active_mass_kg", "deficit_active_mass": "deficit_active_mass_kg", "coverage_percent": "coverage_percent", "raw_mass_used": "raw_mass_used_kg", "loss": "loss_kg"}[metric]
                    series = [{"name": title, "values": [round(sum(float(item.get(key, 0) or 0) for item in comparisons[name]["materials"].values()), 3) for name in labels]}]
            elif metric in {"safe_growth_percent", "deficit_delta", "base_deficit_active_mass", "new_deficit_active_mass", "base_rework_batch_count", "new_rework_batch_count"}:
                changes = args.get("changes_percent") or {}
                if not changes: raise ValueError("Для графика сценария нужны changes_percent")
                scenario = tool("simulate_requirement_change", {"base_requirements": requirements, "changes_percent": changes, "policy": policy})
                labels = sorted(scenario["comparison"])
                key = {"safe_growth_percent": "safe_growth_percent", "deficit_delta": "deficit_delta_active_mass_kg", "base_deficit_active_mass": "base_deficit_active_mass_kg", "new_deficit_active_mass": "new_deficit_active_mass_kg", "base_rework_batch_count": "base_rework_batches", "new_rework_batch_count": "new_rework_batches"}[metric]
                series = [{"name": title, "values": [round(float(scenario["comparison"][name].get(key, 0) or 0), 3) for name in labels]}]
            else:
                plan = build_plan(requirements, policy, True)
                labels, series = _chart_from_plan(plan, metric, group_by, title, unit)
            return {"chart_id": str(uuid.uuid4()), "chart_type": "metric", "metric": metric, "group_by": group_by, "title": title, "labels": labels, "series": series, "unit": unit, "material_type": material, "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], {"metric": metric, "group_by": group_by, "requirements": args.get("requirements"), "policy": args.get("policy") or "hybrid"})}
        rows = [batch_dict(row) for row in con.execute("SELECT * FROM batches" + (" WHERE material_type=?" if material else ""), (material,) if material else ()).fetchall()]; con.close()
        if chart_type == "inventory":
            labels, series = _chart_payload(rows, "raw_mass", "material", material)[0:2]; series.append({"name": "Доступное активное вещество, кг", "values": _chart_payload(rows, "active_mass", "material", material)[1][0]["values"]}); title, unit = "Остатки по материалам", "kg"
        else:
            labels, series, title, unit = _chart_payload(rows, metric, group_by, material)
            if chart_type == "material_quality": title = "Количество партий по материалам и качеству"
            if chart_type == "quality": title = "Распределение по статусам"
        return {"chart_id": str(uuid.uuid4()), "chart_type": chart_type, "metric": metric, "group_by": group_by, "title": title, "labels": labels, "series": series, "unit": unit, "material_type": material, "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], {"material_type": material, "chart_type": chart_type, "metric": metric, "group_by": group_by})}
    if name == "classify_batches":
        material = args.get("material_type"); rows = con.execute("SELECT * FROM batches" + (" WHERE material_type=?" if material else ""), (material,) if material else ()).fetchall(); con.close(); counts = {"GOOD": 0, "REWORK": 0, "REJECTED": 0}
        for row in rows: counts[classify(dict(row))["status"]] += 1
        return {"checked": len(rows), **counts, "run_id": str(uuid.uuid4()), "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], args)}
    if name == "get_inventory_summary":
        material = args.get("material_type"); rows = [batch_dict(r) for r in con.execute("SELECT * FROM batches" + (" WHERE material_type=?" if material else ""), (material,) if material else ()).fetchall()]; con.close(); summary = {}
        for row in rows:
            key = f'{row["material_type"]}:{row["quality"]["status"]}' if args.get("group_by", "material_and_status") == "material_and_status" else row["material_type"]
            item = summary.setdefault(key, {"material_type": row["material_type"], "status": row["quality"]["status"], "batch_count": 0, "raw_mass_kg": 0, "theoretical_active_mass_kg": 0, "available_active_mass_kg": 0, "recovery_loss_active_mass_kg": 0})
            remaining_theoretical_active = row["remaining_raw_mass_kg"] * row["concentration_percent"] / 100
            item["batch_count"] += 1; item["raw_mass_kg"] += row["remaining_raw_mass_kg"]; item["theoretical_active_mass_kg"] += remaining_theoretical_active; item["available_active_mass_kg"] += row["available_active_mass_kg"]
            item["recovery_loss_active_mass_kg"] += remaining_theoretical_active - row["available_active_mass_kg"]
        groups = list(summary.values())
        totals = {key: sum(item[key] for item in groups) for key in ("batch_count", "raw_mass_kg", "theoretical_active_mass_kg", "available_active_mass_kg", "recovery_loss_active_mass_kg")}
        return {"groups": groups, "totals": totals, "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], {"material_type": material, "group_by": args.get("group_by", "material_and_status")})}
    if name == "build_weekly_plan":
        con.close(); return build_plan(args.get("requirements") or requirements_default(), args.get("policy", "hybrid"), args.get("allow_rework", True))
    if name == "check_material_deficit":
        plan = build_plan(args.get("requirements") or {}, "hybrid", args.get("include_rework", True)); return {"materials": plan["materials"], "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], args)}
    if name == "compare_allocation_policies":
        con.close(); req = args.get("requirements") or requirements_default(); return {"policies": {p: build_plan(req, p, True) for p in args.get("policies", POLICIES)}, "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], args)}
    if name == "generate_rejection_report":
        material = args.get("material_type"); rows = [batch_dict(r) for r in con.execute("SELECT * FROM batches" + (" WHERE material_type=?" if material else ""), (material,) if material else ()).fetchall()]; con.close(); statuses = set([x for x, enabled in (("REWORK", args.get("include_rework", True)), ("REJECTED", args.get("include_rejected", True))) if enabled]); selected = [r for r in rows if r["quality"]["status"] in statuses]; return {"report_id": str(uuid.uuid4()), "batches": selected, "total_batches": len(selected), "total_raw_mass_kg": sum(r["remaining_raw_mass_kg"] for r in selected), "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], args)}
    if name == "simulate_requirement_change":
        con.close(); base = args.get("base_requirements") or requirements_default(); changes = args.get("changes_percent") or {}; materials = tuple(sorted(set(material_codes()) | set(base) | set(changes))); new = {m: base.get(m, 0) * (1 + changes.get(m, 0) / 100) for m in materials}; base_plan = build_plan(base, args.get("policy", "hybrid")); new_plan = build_plan(new, args.get("policy", "hybrid")); comparison = {}
        for material in materials:
            before = base_plan["materials"][material]; after = new_plan["materials"][material]; required = float(base.get(material, 0) or 0); safe_growth = max(0.0, (before["available_active_mass_kg"] / required - 1) * 100) if required else 0.0
            comparison[material] = {"base_deficit_active_mass_kg": before["deficit_active_mass_kg"], "new_deficit_active_mass_kg": after["deficit_active_mass_kg"], "deficit_delta_active_mass_kg": round(after["deficit_active_mass_kg"] - before["deficit_active_mass_kg"], 3), "base_rework_batches": sum(item["status"] == "REWORK" for item in before["items"]), "new_rework_batches": sum(item["status"] == "REWORK" for item in after["items"]), "safe_growth_percent": round(safe_growth, 2)}
        return {"base": base_plan, "new": new_plan, "new_requirements": new, "comparison": comparison, "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], args)}
    con.close(); raise ValueError("unknown tool")


TOOLS = tuple(TOOL_REGISTRY)


def tool_specs() -> list[dict[str, Any]]:
    return [{"type": "function", "function": {"name": name, "description": spec["description"], "parameters": {"type": "object", "properties": spec["parameters"], "required": spec["required"], "additionalProperties": False}}} for name, spec in TOOL_REGISTRY.items()]


def result_numbers(data: Any) -> set[float]:
    numbers: set[float] = set()
    if isinstance(data, dict):
        for value in data.values(): numbers.update(result_numbers(value))
    elif isinstance(data, (int, float)) and not isinstance(data, bool):
        numbers.add(round(float(data), 6))
    elif isinstance(data, list):
        for value in data: numbers.update(result_numbers(value))
    return numbers


def answer_numbers_are_grounded(answer: str, data: Any) -> bool:
    if not data: return True
    expected = result_numbers(data)
    numeric_text = re.sub(r"\b\d{4}-\d{2}-\d{2}(?:T\d{2}:\d{2}:\d{2}(?:\.\d+)?Z?)?\b", " ", answer or "")
    numeric_text = re.sub(r"(?<!\w)[A-Za-zА-Яа-яЁё0-9]+(?:-[A-Za-zА-Яа-яЁё0-9]+){1,}(?!\w)", " ", numeric_text)
    for token in re.findall(r"(?<![A-Za-zА-Яа-яЁё0-9-])[-+]?\d+(?:[.,]\d+)?(?![A-Za-zА-Яа-яЁё0-9-])", numeric_text):
        value = float(token.replace(",", "."))
        if not any(abs(value - candidate) <= max(0.051, abs(candidate) * 0.00001) for candidate in expected): return False
    return True


def registry_explanation(tool_name: str) -> str:
    spec = TOOL_REGISTRY[tool_name]
    params = ", ".join(spec["parameters"]) or "без обязательных параметров"
    return f"{spec['title']} ({tool_name}) — {spec['description']} Параметры: {params}. Инструмент read-only: производственные данные не изменяет."


def forced_tool_for_message(message: str, history: list[dict[str, str]] | None = None) -> tuple[str, dict[str, Any]] | None:
    intent = route_intent(message, history)
    if intent["intent"] != "EXECUTE_TOOL": return None
    if intent.get("tool_name") and intent.get("arguments"):
        return intent["tool_name"], intent["arguments"]
    normalized = message.lower()
    material = next((item for item in material_codes() if re.search(rf"\b{re.escape(item.lower())}\b", normalized)), None)
    if any(word in normalized for word in ("остат", "склад", "сырь")):
        return "get_inventory_summary", {"material_type": material, "group_by": "material_and_status"}
    if any(word in normalized for word in ("брак", "отбрак", "доработ")):
        return "generate_rejection_report", {"material_type": material, "include_rework": True, "include_rejected": True}
    batch_match = re.search(r"\b([a-z][a-z0-9_-]{0,15}-\d+)\b", normalized)
    if batch_match and any(word in normalized for word in ("проверь", "качество", "статус", "парт")):
        return "check_batch_quality", {"batch_id": batch_match.group(1).upper()}
    if any(word in normalized for word in ("недельный план", "составь план", "план производства", "построй план")):
        requirements = parse_requirements(message)
        if set(requirements) == set(MATERIALS): return "build_weekly_plan", {"requirements": requirements, "policy": "hybrid", "allow_rework": True}
    return None


def requested_material(message: str) -> str | None:
    normalized = normalize_material_aliases(message.lower())
    if re.search(r"\b(ашки|а-шки)\b", normalized): return "A"
    if re.search(r"\bб\b", normalized): return "B"
    if re.search(r"\bс\b", normalized): return "C"
    return next((item for item in material_codes() if re.search(rf"\b{re.escape(item.lower())}\b", normalized)), None)


def material_choices(prompt: str) -> list[dict[str, str]]:
    return [{"label": "Все материалы", "value": f"{prompt} по всем материалам"}] + [{"label": item, "value": f"{prompt} по материалу {item}"} for item in material_codes()]


def normalize_material_aliases(message: str) -> str:
    return re.sub(r"(?<!\w)а(?!\w)", "a", re.sub(r"(?<!\w)в(?!\w)", "b", re.sub(r"(?<!\w)с(?!\w)", "c", message)))


def parse_requirement_details(message: str) -> tuple[dict[str, float], str | None]:
    normalized = normalize_material_aliases(message.lower()).replace(",", ".")
    values: dict[str, float] = {}
    units = {"т": 1000.0, "тонн": 1000.0, "тонны": 1000.0, "kg": 1.0, "кг": 1.0, "г": 0.001, "g": 0.001}
    found_unit: str | None = None
    for material in material_codes():
        match = re.search(rf"(?:материал\s*)?\b{material.lower()}\b(?:\s+(?:на\s+)?(?:потребность|нужно|нужна|спрос)\s*)?\s*(?:[:=]|[-–])?\s*(\d+(?:\.\d+)?)\s*(т|тонн|тонны|кг|kg|г|g)?", normalized)
        if not match:
            match = re.search(rf"(\d+(?:\.\d+)?)\s*(т|тонн|тонны|кг|kg|г|g)\s*(?:материал\s*)?\b{material.lower()}\b", normalized)
        if match:
            unit = match.group(2) or "кг"
            values[material] = float(match.group(1)) * units[unit]
            found_unit = unit
    return values, found_unit


def parse_requirements(message: str) -> dict[str, float]:
    return parse_requirement_details(message)[0]


def parse_changes(message: str) -> dict[str, float]:
    normalized = message.lower().replace(",", ".")
    changes: dict[str, float] = {}
    for material in material_codes():
        match = re.search(rf"\b{material.lower()}\b[^%]{{0,45}}?(?:на|рост\w*|выраст\w*|увелич\w*)\s*([+-]?\d+(?:\.\d+)?)\s*%", normalized)
        if match: changes[material] = float(match.group(1))
    return changes


def mass_basis(message: str) -> str | None:
    normalized = message.lower()
    if any(word in normalized for word in ("активн", "действующ", "active")):
        return "active_mass_kg"
    if any(word in normalized for word in ("сыр", "raw")):
        return "raw_mass_kg"
    return None


def clarification_text(intent: dict[str, Any]) -> str:
    missing = intent.get("missing_fields", [])
    if "mass_basis" in missing:
        return "Уточните, числа указаны как масса активного вещества или как масса сырья."
    if intent.get("tool_name") == "check_material_deficit" and "requirements" in missing:
        return "Укажите потребность в кг активного вещества, например: A 3000, B 2500, C 1800."
    return "Уточните, пожалуйста: " + ", ".join(missing or ["параметры запроса"]) + "."


def chart_request(message: str, material: str | None) -> dict[str, Any]:
    normalized = message.lower()
    quality = any(word in normalized for word in ("качеств", "статус"))
    if any(word in normalized for word in ("брак", "отбрак", "доработ")):
        metric = "rejection_share" if "дол" in normalized or "процент" in normalized else "rejection_active_mass" if "активн" in normalized else "rejection_raw_mass" if any(word in normalized for word in ("масс", "кг", "остат")) else "rejection_batch_count"
    elif any(word in normalized for word in ("сценар", "изменен", "рост потребн", "увелич")):
        metric = "safe_growth_percent" if any(word in normalized for word in ("безопас", "рост")) else "deficit_delta"
    elif any(word in normalized for word in ("дефицит", "нехват", "не хват")):
        metric = "deficit_active_mass"
    elif any(word in normalized for word in ("покрыт", "обеспечен")):
        metric = "coverage_percent"
    elif any(word in normalized for word in ("потер", "восстановлен")):
        metric = "loss" if "план" in normalized else "recovery_loss"
    elif any(word in normalized for word in ("требуем", "потребн")):
        metric = "required_active_mass"
    elif "доступн" in normalized and "план" in normalized:
        metric = "plan_available_active_mass"
    elif any(word in normalized for word in ("выбранн", "отобранн")) or ("использован" in normalized and "парти" in normalized):
        metric = "planned_batch_count"
    elif "теорет" in normalized:
        metric = "theoretical_active_mass"
    elif any(word in normalized for word in ("активн", "действующ")):
        metric = "active_mass"
    elif "концентрац" in normalized:
        metric = "concentration"
    elif any(word in normalized for word in ("возраст", "старост", "старые", "старейш")):
        metric = "age_days"
    elif any(word in normalized for word in ("количеств", "сколько партий", "число партий")):
        metric = "status_count" if quality else "batch_count"
    elif "дол" in normalized and quality:
        metric = "status_share"
    elif quality:
        metric = "status_count"
    else:
        metric = "raw_mass"
    group_by = "status" if "по статус" in normalized else "batch" if any(word in normalized for word in ("по партиям", "каждой партии")) else "policy" if any(word in normalized for word in ("стратег", "политик", "fifo", "hybrid", "сравни")) else "material"
    if group_by == "policy" and metric == "raw_mass": metric = "coverage_percent"
    chart_type = "material_quality" if metric == "status_count" and group_by == "material" else "quality" if metric == "status_count" and group_by == "status" else "concentration" if metric == "concentration" and group_by == "material" else "metric"
    return {"chart_type": chart_type, "metric": metric, "group_by": group_by, "material_type": material}


def route_intent(message: str, history: list[dict[str, str]] | None = None) -> dict[str, Any]:
    normalized = message.lower()
    explanation = explanation_tool_for_message(message)
    if explanation:
        return {"intent": "EXPLAIN_TOOL", "tool_name": explanation, "arguments": {}, "missing_fields": [], "confidence": 0.99, "reason": "Запрос содержит просьбу объяснить инструмент"}
    if any(word in normalized for word in ("открой обзор", "открой партии", "открой план", "открой отчёт", "открой отчет", "открой ассистент")):
        target = "data" if "парт" in normalized else "planning" if "план" in normalized else "reports" if "отч" in normalized else "assistant" if "ассист" in normalized else "overview"
        return {"intent": "NAVIGATE", "tool_name": None, "arguments": {"page": target}, "missing_fields": [], "confidence": 0.98, "reason": "Явная команда навигации"}
    context = "\n".join(item.get("content", "") for item in (history or [])[-8:]).lower()
    previous_user = next((item.get("content", "") for item in reversed(history or []) if item.get("role") == "user"), "")
    previous_requirements = parse_requirements(previous_user)
    followup_requirements = parse_requirements(message)
    if history and followup_requirements and any(word in context for word in ("дефицит", "потребност", "укажите потребность")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "check_material_deficit", "arguments": {"requirements": followup_requirements, "include_rework": True, "mass_basis": "active_mass_kg"}, "missing_fields": [], "confidence": 0.94, "reason": "Потребность уточнена коротким follow-up после запроса дефицита"}
    if history and requested_material(message) and any(word in normalized for word in ("а теперь", "теперь", "ещё", "еще")) and any(word in context for word in ("остат", "склад", "запас", "inventory")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "get_inventory_summary", "arguments": {"material_type": requested_material(message), "group_by": "material_and_status"}, "missing_fields": [], "confidence": 0.91, "reason": "Материал уточнён по контексту предыдущего запроса"}
    if history and previous_requirements and set(previous_requirements) == set(MATERIALS) and any(word in normalized for word in ("такой же", "как до этого", "как раньше", "но fifo", "но hybrid", "тоже")) and "план" in context:
        policy = "strict_fifo" if "fifo" in normalized else "max_concentration" if "концентрац" in normalized else "hybrid"
        return {"intent": "EXECUTE_TOOL", "tool_name": "build_weekly_plan", "arguments": {"requirements": previous_requirements, "policy": policy, "allow_rework": True, "mass_basis": mass_basis(previous_user) or "active_mass_kg"}, "missing_fields": [], "confidence": 0.9, "reason": "Параметры плана унаследованы из контекста"}
    if history and previous_requirements and "план" in context and any(word in normalized for word in ("%", "процент", "больше", "увеличь", "рост")):
        changes = {material.upper(): float(value) for material, value in re.findall(r"\b([a-z][a-z0-9_-]{0,15})\b\s*(?:на\s*)?([+-]?\d+(?:[.,]\d+)?)\s*%", normalized) if material.upper() in material_codes()}
        policy = "strict_fifo" if "fifo" in context else "max_concentration" if "концентрац" in context else "hybrid"
        if changes:
            return {"intent": "EXECUTE_TOOL", "tool_name": "simulate_requirement_change", "arguments": {"base_requirements": previous_requirements, "changes_percent": changes, "policy": policy}, "missing_fields": [], "confidence": 0.88, "reason": "Сценарий изменения унаследовал базовую потребность из контекста"}
    material = requested_material(message)
    batch_match = re.search(r"\b([a-z][a-z0-9_-]{0,15}-\d+)\b", normalized)
    if batch_match and any(word in normalized for word in ("детал", "покажи", "карточ")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "get_batch_details", "arguments": {"batch_id": batch_match.group(1).upper()}, "missing_fields": [], "confidence": 0.99, "reason": "Явно запрошены детали партии"}
    if batch_match and any(word in normalized for word in ("проверь", "качество", "статус", "парт")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "check_batch_quality", "arguments": {"batch_id": batch_match.group(1).upper()}, "missing_fields": [], "confidence": 0.99, "reason": "Явно указана партия и команда проверки"}
    if any(word in normalized for word in ("проверь качество партии", "проверь партию", "статус партии", "карточка партии")):
        con = db(); choices = [row[0] for row in con.execute("SELECT batch_id FROM batches ORDER BY arrival_date, batch_id LIMIT 20")]; con.close()
        return {"intent": "CLARIFY", "tool_name": "check_batch_quality", "arguments": {}, "missing_fields": ["batch_id"], "confidence": 0.96, "reason": "Нужно выбрать партию для проверки", "choices": choices}
    if any(word in normalized for word in ("график", "диаграмм", "визуализ", "построй граф")) and not any(word in normalized for word in ("брак", "отбрак", "отчёт", "отчет", "отклон")):
        arguments = chart_request(message, material)
        if arguments["metric"] in {"deficit_active_mass", "coverage_percent", "loss"} or arguments["group_by"] == "policy":
            requirements = parse_requirements(message)
            if not requirements:
                return {"intent": "CLARIFY", "tool_name": "build_chart", "arguments": arguments, "missing_fields": ["requirements"], "confidence": 0.92, "reason": "Для графика плана нужна потребность по активному веществу"}
            arguments["requirements"] = requirements
        if arguments["metric"] in {"safe_growth_percent", "deficit_delta"}:
            changes = parse_changes(message)
            if not changes:
                return {"intent": "CLARIFY", "tool_name": "build_chart", "arguments": arguments, "missing_fields": ["changes_percent"], "confidence": 0.92, "reason": "Для графика сценария нужен процент изменения потребности"}
            arguments["changes_percent"] = changes
            arguments["requirements"] = parse_requirements(message) or requirements_default()
            arguments["policy"] = "strict_fifo" if "fifo" in normalized else "max_concentration" if "концентрац" in normalized else "hybrid"
        return {"intent": "EXECUTE_TOOL", "tool_name": "build_chart", "arguments": arguments, "missing_fields": [], "confidence": 0.98, "reason": "Запрошена визуализация доступной метрики"}
    if any(word in normalized for word in ("самую стар", "самые стар", "старейш", "ранние партии")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "get_oldest_batches", "arguments": {"material_type": material, "limit": 5}, "missing_fields": [], "confidence": 0.98, "reason": "Запрошены партии с самой ранней датой поступления"}
    if any(word in normalized for word in ("классифиц", "классификац")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "classify_batches", "arguments": {"material_type": material, "only_unclassified": False}, "missing_fields": [], "confidence": 0.96, "reason": "Явно запрошена классификация партий"}
    if any(word in normalized for word in ("остат", "остал", "склад", "запас", "сырь", "че по", "потер")) and any(word in normalized for word in ("покажи", "покаж", "дай", "проверь", "посчитай", "сколько", "че по", "где")):
        if not material and not any(word in normalized for word in ("все", "всем", "общ", "где")):
            return {"intent": "CLARIFY", "tool_name": "get_inventory_summary", "arguments": {}, "missing_fields": ["material_type"], "confidence": 0.95, "reason": "Нужно выбрать материал или все материалы", "choices": material_choices("Покажи остатки")}
        return {"intent": "EXECUTE_TOOL", "tool_name": "get_inventory_summary", "arguments": {"material_type": material, "group_by": "material_and_status"}, "missing_fields": [], "confidence": 0.98, "reason": "Явно запрошена сводка остатков"}
    if any(word in normalized for word in ("график", "диаграмм", "визуализ", "построй граф")) and any(word in normalized for word in ("брак", "отбрак", "доработ")) and not any(word in normalized for word in ("отчёт", "отчет", "сформируй отчёт", "сформируй отчет")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "build_chart", "arguments": chart_request(message, material), "missing_fields": [], "confidence": 0.98, "reason": "Запрошена визуализация брака"}
    if any(word in normalized for word in ("брак", "отбрак", "доработ", "проблемн", "отклон")) and any(word in normalized for word in ("покажи", "сделай", "сформируй", "отчёт", "отчет", "парти")):
        if not material and not any(word in normalized for word in ("все", "всем")):
            prompt = "Создай отчёт по браку и построй график брака" if any(word in normalized for word in ("график", "диаграмм", "визуализ")) else "Покажи отчёт по браку"
            return {"intent": "CLARIFY", "tool_name": "generate_rejection_report", "arguments": {}, "missing_fields": ["material_type"], "confidence": 0.95, "reason": "Нужно выбрать материал или все материалы", "choices": material_choices(prompt)}
        return {"intent": "EXECUTE_TOOL", "tool_name": "generate_rejection_report", "arguments": {"material_type": material, "include_rework": True, "include_rejected": True}, "missing_fields": [], "confidence": 0.98, "reason": "Явно запрошен отчёт по браку"}
    if any(word in normalized for word in ("недельный план", "составь план", "построй план", "план производства")):
        requirements = parse_requirements(message); missing = [m for m in MATERIALS if m not in requirements]
        policy = "strict_fifo" if "fifo" in normalized and "hybrid" not in normalized else "max_concentration" if any(alias in normalized for alias in ("концентрац", "max concentration", "max_concentration")) else "hybrid" if "hybrid" in normalized else None
        missing_fields = [f"requirements.{m}" for m in missing] + ([] if policy else ["policy"])
        if not missing and not mass_basis(message): missing_fields.append("mass_basis")
        if missing_fields:
            return {"intent": "CLARIFY", "tool_name": "build_weekly_plan", "arguments": {"requirements": requirements}, "missing_fields": missing_fields, "confidence": 0.94, "reason": "Для preview-плана не хватает обязательных вводных", "choices": [{"label": "Hybrid", "value": "Построй недельный план с policy hybrid"}, {"label": "Strict FIFO", "value": "Построй недельный план с policy strict_fifo"}, {"label": "Max concentration", "value": "Построй недельный план с policy max_concentration"}]}
        return {"intent": "EXECUTE_TOOL", "tool_name": "build_weekly_plan", "arguments": {"requirements": requirements, "policy": policy, "allow_rework": True, "mass_basis": mass_basis(message)}, "missing_fields": [], "confidence": 0.98, "reason": "Все обязательные параметры preview-плана указаны"}
    if any(word in normalized for word in ("дефицит", "хватит", "потребн", "вытян", "достат")) and any(word in normalized for word in ("проверь", "посчитай", "есть", "хватит", "вытян", "достат")):
        requirements = parse_requirements(message)
        if not requirements:
            if any(word in normalized for word in ("текущ", "сохранён", "сохранен", "сохраненну")):
                return {"intent": "EXECUTE_TOOL", "tool_name": "check_material_deficit", "arguments": {"requirements": requirements_default(), "include_rework": True}, "missing_fields": [], "confidence": 0.94, "reason": "Пользователь явно выбрал сохранённую потребность"}
            return {"intent": "CLARIFY", "tool_name": "check_material_deficit", "arguments": {}, "missing_fields": ["requirements"], "confidence": 0.95, "reason": "Для проверки дефицита нужна явная потребность", "choices": [{"label": "Использовать сохранённую потребность", "value": "Проверь дефицит по текущей потребности"}]}
        if not mass_basis(message): return {"intent": "CLARIFY", "tool_name": "check_material_deficit", "arguments": {"requirements": requirements}, "missing_fields": ["mass_basis"], "confidence": 0.94, "reason": "Нужно различить массу сырья и активного вещества"}
        return {"intent": "EXECUTE_TOOL", "tool_name": "check_material_deficit", "arguments": {"requirements": requirements, "include_rework": True, "mass_basis": mass_basis(message)}, "missing_fields": [], "confidence": 0.96, "reason": "Указана потребность для проверки дефицита"}
    if any(word in normalized for word in ("сравни", "стратег", "политик")) and any(word in normalized for word in ("fifo", "hybrid", "концентрац", "стратег")):
        requirements = parse_requirements(message)
        if not requirements: return {"intent": "CLARIFY", "tool_name": "compare_allocation_policies", "arguments": {}, "missing_fields": ["requirements"], "confidence": 0.9, "reason": "Нужна потребность для честного сравнения стратегий"}
        if not mass_basis(message): return {"intent": "CLARIFY", "tool_name": "compare_allocation_policies", "arguments": {"requirements": requirements}, "missing_fields": ["mass_basis"], "confidence": 0.94, "reason": "Нужно различить массу сырья и активного вещества"}
        policies = [p for p, aliases in (("strict_fifo", ("fifo", "strict_fifo")), ("hybrid", ("hybrid",)), ("max_concentration", ("концентрац", "max_concentration"))) if any(alias in normalized for alias in aliases)] or list(POLICIES)
        return {"intent": "EXECUTE_TOOL", "tool_name": "compare_allocation_policies", "arguments": {"requirements": requirements, "policies": policies, "mass_basis": mass_basis(message)}, "missing_fields": [], "confidence": 0.96, "reason": "Указаны потребность и стратегии"}
    if any(word in normalized for word in ("сценар", "смоделируй", "изменение спроса", "что будет, если", "рост")):
        changes = parse_changes(message)
        policy = "strict_fifo" if "fifo" in normalized and "hybrid" not in normalized else "max_concentration" if any(alias in normalized for alias in ("концентрац", "max concentration", "max_concentration")) else "hybrid" if "hybrid" in normalized else None
        missing_fields = ([] if changes else ["changes_percent"]) + ([] if policy else ["policy"])
        if missing_fields:
            return {"intent": "CLARIFY", "tool_name": "simulate_requirement_change", "arguments": {"changes_percent": changes}, "missing_fields": missing_fields, "confidence": 0.92, "reason": "Для сценария не хватает процента изменения и политики", "choices": [{"label": "Hybrid", "value": "Смоделируй сценарий с policy hybrid"}, {"label": "Strict FIFO", "value": "Смоделируй сценарий с policy strict_fifo"}]}
        return {"intent": "EXECUTE_TOOL", "tool_name": "simulate_requirement_change", "arguments": {"base_requirements": {}, "changes_percent": changes, "policy": policy}, "missing_fields": [], "confidence": 0.95, "reason": "Указаны параметры сценария"}
    if any(word in normalized for word in ("привет", "что ты умеешь", "помоги", "что можешь", "как пользоваться")):
        return {"intent": "GENERAL_HELP", "tool_name": None, "arguments": {}, "missing_fields": [], "confidence": 0.98, "reason": "Общий вопрос без бизнес-команды"}
    return {"intent": "GENERAL_HELP", "tool_name": None, "arguments": {}, "missing_fields": [], "confidence": 0.55, "reason": "Не найдено однозначной команды на выполнение"}


def explanation_tool_for_message(message: str) -> str | None:
    normalized = message.lower()
    if not any(word in normalized for word in ("как работает", "как устроен", "объясни", "расскажи про", "что делает", "зачем нужен", "какие параметры", "как рассчитывается", "не запускай", "только объясни")):
        return None
    for name, spec in TOOL_REGISTRY.items():
        aliases = (spec["title"].lower(), name.replace("_", " "), *spec.get("aliases", []))
        if any(alias in normalized for alias in aliases): return name
    return None


def compact_history(history: list[dict[str, str]]) -> list[dict[str, str]]:
    """Keep only the small conversational window needed for an ambiguous request."""
    compact = []
    for item in history[-6:]:
        content = (item.get("content") or "").strip()
        if content:
            compact.append({"role": item["role"], "content": content[:700]})
    return compact


def llm_context(history: list[dict[str, str]]) -> str:
    last_user = next((item["content"] for item in reversed(history) if item.get("role") == "user"), "")
    return json.dumps({
        "last_user_request": last_user[:700],
        "last_requirements": parse_requirements(last_user),
        "last_material": requested_material(last_user),
        "last_policy": "strict_fifo" if "fifo" in last_user.lower() else "max_concentration" if "концентрац" in last_user.lower() else "hybrid" if "hybrid" in last_user.lower() else None,
    }, ensure_ascii=False)


def compound_request(message: str, history: list[dict[str, str]]) -> list[tuple[str, dict[str, Any]]] | None:
    """Resolve safe read-only workflows that intentionally combine tools."""
    normalized = message.lower()
    intent = route_intent(message, history)
    if intent["intent"] != "EXECUTE_TOOL" or not intent.get("arguments"):
        return None
    args = intent["arguments"]
    material = args.get("material_type")
    if any(word in normalized for word in ("график", "диаграм", "визуализ")) and any(word in normalized for word in ("остат", "склад", "запас")):
        return [("get_inventory_summary", {"material_type": material, "group_by": "material_and_status"}), ("build_chart", chart_request(message, material))]
    if any(word in normalized for word in ("график", "диаграм", "визуализ")) and any(word in normalized for word in ("брак", "отбрак", "отчёт", "отчет", "отклон")):
        chart_args = chart_request(message, material)
        if chart_args["metric"] == "rejection_batch_count" and not any(word in normalized for word in ("колич", "сколько", "число")):
            chart_args["chart_type"] = "quality"
        return [("generate_rejection_report", args), ("build_chart", chart_args)]
    if any(word in normalized for word in ("классифиц", "классификац")) and any(word in normalized for word in ("брак", "отчёт", "отчет", "отклон")):
        return [("classify_batches", args), ("generate_rejection_report", {"material_type": material, "include_rework": True, "include_rejected": True})]
    if any(word in normalized for word in ("проверь качество", "статус партии")) and any(word in normalized for word in ("детал", "карточ")) and args.get("batch_id"):
        return [("check_batch_quality", {"batch_id": args["batch_id"]}), ("get_batch_details", {"batch_id": args["batch_id"]})]
    if "план" in normalized and any(word in normalized for word in ("дефицит", "стратег", "политик")) and args.get("requirements"):
        if any(word in normalized for word in ("дефицит", "хватит")):
            return [("build_weekly_plan", args), ("check_material_deficit", {"requirements": args["requirements"], "include_rework": args.get("allow_rework", True)})]
        policies = [p for p, aliases in (("strict_fifo", ("fifo",)), ("hybrid", ("hybrid",)), ("max_concentration", ("концентрац", "max concentration"))) if any(alias in normalized for alias in aliases)] or list(POLICIES)
        return [("build_weekly_plan", args), ("compare_allocation_policies", {"requirements": args["requirements"], "policies": policies})]
    if "план" in normalized and any(word in normalized for word in ("график", "диаграм", "визуализ")) and args.get("requirements"):
        chart_args = chart_request(message, material); chart_args.update({"requirements": args["requirements"], "policy": args.get("policy", "hybrid")})
        return [("build_weekly_plan", args), ("build_chart", chart_args)]
    return None


def routed_tool_result(message: str, history: list[dict[str, str]]) -> tuple[str, Any, list[dict[str, Any]]] | None:
    intent = route_intent(message, history)
    if intent["intent"] != "EXECUTE_TOOL" or not intent.get("tool_name") or not intent.get("arguments"):
        return None
    name, args = intent["tool_name"], intent["arguments"]
    try:
        compound = compound_request(message, history)
        if compound:
            results = {}
            trace = []
            for call_name, call_args in compound:
                results[call_name] = tool(call_name, call_args)
                trace.append({"tool": call_name, "arguments": call_args, "status": "success", "source": "router"})
            data = {"results": results}
            if "generate_rejection_report" in results: data["report"] = results["generate_rejection_report"]
            if "build_chart" in results: data["chart"] = results["build_chart"]
            return compound[0][0], data, trace
        if name == "generate_rejection_report" and any(word in message.lower() for word in ("график", "диаграмм", "визуализ")):
            report = tool(name, args)
            chart_args = chart_request(message, args.get("material_type"))
            if chart_args["metric"] == "rejection_batch_count": chart_args["chart_type"] = "quality"
            chart = tool("build_chart", chart_args)
            trace = [
                {"tool": name, "arguments": args, "status": "success", "source": "router"},
                {"tool": "build_chart", "arguments": chart_args, "status": "success", "source": "router"},
            ]
            return name, {"report": report, "chart": chart}, trace
        data = tool(name, args)
        trace = [{"tool": name, "arguments": args, "status": "success", "source": "router"}]
    except Exception as exc:
        data = {"error": str(exc)}
        trace = [{"tool": name, "arguments": args, "status": "error", "message": str(exc), "source": "router"}]
    return name, data, trace


def explanation_prompt(message: str, history: list[dict[str, str]], tool_name: str | None = None, data: Any = None) -> tuple[str, list[dict[str, str]]]:
    system = """Ты — AI-технолог системы контроля качества сырья. Отвечай по-русски, коротко и по делу. Не придумывай числа, партии или статусы. Если ниже есть результат инструмента, объясни именно его простыми словами и отделяй массу сырья от массы активного вещества. Используй только числа, которые явно есть в JSON результата; новые суммы, проценты и производные показатели не рассчитывай. Не запускай инструменты и не меняй данные."""
    if tool_name and data is not None:
        prompt_data = data
        if tool_name == "generate_rejection_report" and isinstance(data, dict) and isinstance(data.get("batches"), list):
            prompt_data = {**data, "batches": data["batches"][:12], "_note": f"В отчёте всего партий: {len(data['batches'])}. Полный список уже приложен интерфейсу."}
        prompt = f"Запрос пользователя: {message}\nИнструмент {tool_name} уже выполнен. Его результат JSON:\n{json.dumps(prompt_data, ensure_ascii=False, separators=(',', ':'))[:24000]}"
    else:
        prompt = message
    messages = [{"role": "system", "content": system + "\nСжатый контекст: " + llm_context(history)}]
    messages.extend(compact_history(history))
    messages.append({"role": "user", "content": prompt})
    return system, messages


def llm_may_select_tool(message: str) -> bool:
    normalized = message.lower()
    if explanation_tool_for_message(message) or any(word in normalized for word in ("как работает", "как устроен", "объясни", "расскажи", "зачем нужен", "игнорируй", "поставь", "без нового запуска", "почему ты выбрал", "просто объясни")):
        return False
    return any(word in normalized for word in ("остат", "запас", "сырь", "парт", "брак", "план", "дефиц", "стратег", "потребн", "потер"))
