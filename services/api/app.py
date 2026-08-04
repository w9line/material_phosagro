from __future__ import annotations

import csv
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import sqlite3
import uuid
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.request import Request as URLRequest, urlopen

from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field, field_validator

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./data/app.db")
DB_PATH = DATABASE_URL.replace("sqlite:///", "", 1)
if not DATABASE_URL.startswith("postgres"):
    if not DB_PATH.startswith("/"):
        DB_PATH = str(Path(__file__).parent / DB_PATH)
    Path(DB_PATH).parent.mkdir(parents=True, exist_ok=True)

MATERIALS = ("A", "B", "C")
POLICIES = ("strict_fifo", "max_concentration", "hybrid")
DEFAULT_RULES = {
    "A": (28.0, 23.0, 1.0, 0.9, 0.0),
    "B": (30.0, 25.0, 1.0, 0.9, 0.0),
    "C": (35.0, 25.0, 1.0, 0.9, 0.0),
}

TOOL_REGISTRY: dict[str, dict[str, Any]] = {
    "check_batch_quality": {"title": "Качество партии", "aliases": ["проверка партии", "статус партии", "check batch quality"], "description": "Проверяет качество и статус одной партии.", "parameters": {"batch_id": {"type": "string"}}, "required": ["batch_id"], "examples": ["Проверь A-001"], "mutating": False, "category": "quality", "units": {"mass": "kg_raw"}},
    "get_batch_details": {"title": "Детали партии", "aliases": ["параметры партии", "batch details"], "description": "Возвращает фактические параметры партии.", "parameters": {"batch_id": {"type": "string"}}, "required": ["batch_id"], "examples": ["Покажи детали A-001"], "mutating": False, "category": "quality", "units": {"mass": "kg_raw"}},
    "classify_batches": {"title": "Классификация", "aliases": ["классификация партий", "classify batches"], "description": "Классифицирует партии по действующим порогам качества.", "parameters": {"material_type": {"type": ["string", "null"]}, "only_unclassified": {"type": "boolean"}}, "required": [], "examples": ["Классифицируй партии A"], "mutating": False, "category": "quality", "units": {}, "filters": ["material_type"]},
    "get_inventory_summary": {"title": "Остатки", "aliases": ["склад", "запасы", "сырьё", "inventory"], "description": "Показывает остатки сырой массы и доступного активного вещества.", "parameters": {"material_type": {"type": ["string", "null"]}, "group_by": {"type": "string", "enum": ["material_and_status"]}}, "required": [], "examples": ["Покажи остатки по A"], "mutating": False, "category": "inventory", "units": {"raw_mass": "kg_raw", "active_mass": "kg_active"}, "filters": ["material_type"]},
    "build_weekly_plan": {"title": "Недельный план", "aliases": ["план производства", "weekly plan", "планирование"], "description": "Строит безопасный preview-план без списания остатков.", "parameters": {"requirements": {"type": "object"}, "policy": {"type": "string", "enum": list(POLICIES)}, "allow_rework": {"type": "boolean"}}, "required": ["requirements", "policy"], "examples": ["Построй план A 3000 B 2500 C 1800, hybrid"], "mutating": False, "category": "planning", "units": {"requirements": "kg_active"}},
    "check_material_deficit": {"title": "Дефицит", "aliases": ["не хватит", "потребность", "material deficit"], "description": "Сравнивает потребность с доступным активным веществом.", "parameters": {"requirements": {"type": "object"}, "include_rework": {"type": "boolean"}}, "required": ["requirements"], "examples": ["Проверь дефицит A 3000"], "mutating": False, "category": "planning", "units": {"requirements": "kg_active", "deficit": "kg_active"}},
    "compare_allocation_policies": {"title": "Стратегии", "aliases": ["политики распределения", "сравнение стратегий", "compare policies", "allocation policies", "fifo"], "description": "Сравнивает FIFO, max concentration и hybrid без изменения данных.", "parameters": {"requirements": {"type": "object"}, "policies": {"type": "array", "items": {"type": "string", "enum": list(POLICIES)}}}, "required": ["requirements", "policies"], "examples": ["Сравни FIFO и hybrid для A 3000"], "mutating": False, "category": "planning", "units": {"requirements": "kg_active"}},
    "generate_rejection_report": {"title": "Отчёт по браку", "aliases": ["отчёт по отклонениям", "rejection report", "брак"], "description": "Формирует read-only отчёт по REWORK и REJECTED.", "parameters": {"material_type": {"type": ["string", "null"]}, "include_rework": {"type": "boolean"}, "include_rejected": {"type": "boolean"}}, "required": [], "examples": ["Сделай отчёт по браку A"], "mutating": False, "category": "reports", "units": {"mass": "kg_raw"}, "filters": ["material_type"]},
    "simulate_requirement_change": {"title": "Сценарий потребности", "aliases": ["изменение спроса", "simulate requirement", "сценарий"], "description": "Сравнивает базовый и изменённый спрос без изменения склада.", "parameters": {"base_requirements": {"type": "object"}, "changes_percent": {"type": "object"}, "policy": {"type": "string", "enum": list(POLICIES)}}, "required": ["changes_percent", "policy"], "examples": ["Что будет, если B вырастет на 20%"], "mutating": False, "category": "planning", "units": {"requirements": "kg_active", "changes": "percent"}},
}


class CompatRow(dict):
    def __getitem__(self, key: Any) -> Any:
        return list(self.values())[key] if isinstance(key, int) else super().__getitem__(key)


class PostgresCursor:
    def __init__(self, cursor: Any):
        self.cursor = cursor

    @property
    def rowcount(self) -> int:
        return self.cursor.rowcount

    def _row(self, row: Any) -> Any:
        if row is None or not self.cursor.description:
            return row
        return CompatRow(zip([column.name for column in self.cursor.description], row))

    def fetchone(self) -> Any:
        return self._row(self.cursor.fetchone())

    def fetchall(self) -> list[Any]:
        return [self._row(row) for row in self.cursor.fetchall()]

    def __iter__(self):
        for row in self.cursor:
            yield self._row(row)


class PostgresConnection:
    def __init__(self, url: str):
        import psycopg
        self.con = psycopg.connect(url)

    def execute(self, sql: str, params: tuple[Any, ...] = ()) -> Any:
        cursor = self.con.cursor(); cursor.execute(sql.replace("?", "%s"), params); return PostgresCursor(cursor)

    def executemany(self, sql: str, params: Any) -> Any:
        cursor = self.con.cursor(); cursor.executemany(sql.replace("?", "%s"), params); return PostgresCursor(cursor)

    def executescript(self, script: str) -> None:
        for statement in script.split(";"):
            if statement.strip():
                self.execute(statement)

    def commit(self) -> None:
        self.con.commit()

    def rollback(self) -> None:
        self.con.rollback()

    def close(self) -> None:
        self.con.close()


def db() -> Any:
    if DATABASE_URL.startswith("postgres"):
        return PostgresConnection(DATABASE_URL)
    con = sqlite3.connect(DB_PATH, check_same_thread=False)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys=ON")
    return con


def bump_data_version(con: Any) -> int:
    row = con.execute("SELECT value FROM settings WHERE key='data_version'").fetchone()
    version = int(row[0] if row else 0) + 1
    con.execute("INSERT INTO settings(key, value) VALUES ('data_version', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (str(version),))
    return version


def snapshot_meta(units: dict[str, str], parameters: dict[str, Any] | None = None) -> dict[str, Any]:
    con = db(); row = con.execute("SELECT value FROM settings WHERE key='data_version'").fetchone(); con.close()
    return {"data_version": int(row[0]) if row else 1, "calculated_at": datetime.utcnow().isoformat() + "Z", "parameters": parameters or {}, "units": units}


def init_db() -> None:
    con = db()
    con.executescript("""
    CREATE TABLE IF NOT EXISTS batches (
      batch_id TEXT PRIMARY KEY, material_type TEXT NOT NULL, raw_mass_kg REAL NOT NULL,
      concentration_percent REAL NOT NULL, arrival_date TEXT NOT NULL, supplier TEXT,
      warehouse TEXT, certificate_number TEXT, notes TEXT, remaining_raw_mass_kg REAL NOT NULL,
      source TEXT NOT NULL, status TEXT, created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS quality_rules (
      material_type TEXT PRIMARY KEY, good_threshold_percent REAL NOT NULL,
      rework_threshold_percent REAL NOT NULL, good_recovery_factor REAL NOT NULL,
      rework_recovery_factor REAL NOT NULL, reject_recovery_factor REAL NOT NULL
    );
    CREATE TABLE IF NOT EXISTS requirements (material_type TEXT PRIMARY KEY, required_active_mass_kg REAL NOT NULL);
    CREATE TABLE IF NOT EXISTS plans (plan_id TEXT PRIMARY KEY, status TEXT NOT NULL, payload TEXT NOT NULL, created_at TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS users (
      user_id TEXT PRIMARY KEY, username TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
      is_admin INTEGER NOT NULL DEFAULT 0, is_blocked INTEGER NOT NULL DEFAULT 0, created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS sessions (
      token_hash TEXT PRIMARY KEY, user_id TEXT NOT NULL, expires_at TEXT NOT NULL,
      FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS plan_owners (plan_id TEXT PRIMARY KEY, user_id TEXT NOT NULL, status TEXT NOT NULL DEFAULT 'preview', created_at TEXT NOT NULL, preview_data_version INTEGER NOT NULL DEFAULT 1, FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE);
    CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS chats (
      chat_id TEXT PRIMARY KEY, user_id TEXT NOT NULL, title TEXT NOT NULL,
      created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
      FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS chat_messages (
      message_id TEXT PRIMARY KEY, chat_id TEXT NOT NULL, role TEXT NOT NULL,
      content TEXT NOT NULL, tool_calls TEXT NOT NULL DEFAULT '[]', created_at TEXT NOT NULL,
      FOREIGN KEY(chat_id) REFERENCES chats(chat_id) ON DELETE CASCADE
    );
    INSERT INTO settings(key, value) VALUES ('registration_open', '1') ON CONFLICT(key) DO NOTHING;
    INSERT INTO settings(key, value) VALUES ('data_version', '1') ON CONFLICT(key) DO NOTHING;
    """)
    if DATABASE_URL.startswith("postgres"):
        column_exists = con.execute("SELECT 1 FROM information_schema.columns WHERE table_name='plan_owners' AND column_name='preview_data_version'").fetchone()
    else:
        column_exists = next((row for row in con.execute("PRAGMA table_info(plan_owners)").fetchall() if row[1] == "preview_data_version"), None)
    if not column_exists:
        con.execute("ALTER TABLE plan_owners ADD COLUMN preview_data_version INTEGER NOT NULL DEFAULT 1")
    for material, values in DEFAULT_RULES.items():
        con.execute("INSERT INTO quality_rules VALUES (?,?,?,?,?,?) ON CONFLICT(material_type) DO NOTHING", (material, *values))
        con.execute("INSERT INTO requirements VALUES (?,?) ON CONFLICT(material_type) DO NOTHING", (material, 3000.0))
    if con.execute("SELECT COUNT(*) FROM batches").fetchone()[0] == 0:
        demo = [
            ("A-001", "A", 5000, 24.5, "2026-08-01", "Поставщик-1", "Основной", None, "Демо", 5000, "demo", None, datetime.utcnow().isoformat()),
            ("A-002", "A", 4200, 30.1, "2026-08-03", "Поставщик-2", "Основной", None, "Демо", 4200, "demo", None, datetime.utcnow().isoformat()),
            ("B-001", "B", 7000, 38.2, "2026-08-02", "Поставщик-1", "Зона B", None, "Демо", 7000, "demo", None, datetime.utcnow().isoformat()),
            ("C-001", "C", 6000, 32.0, "2026-08-04", "Поставщик-3", "Зона C", None, "Демо", 6000, "demo", None, datetime.utcnow().isoformat()),
        ]
        con.executemany("INSERT INTO batches VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", demo)
    admin_username = os.getenv("ADMIN_USERNAME", "admin").strip()
    admin_password = os.getenv("ADMIN_PASSWORD", "")
    if admin_password and not con.execute("SELECT 1 FROM users WHERE username=?", (admin_username,)).fetchone():
        con.execute("INSERT INTO users VALUES (?,?,?,?,?,?)", (str(uuid.uuid4()), admin_username, hash_password(admin_password), 1, 0, datetime.utcnow().isoformat()))
    con.commit()
    con.close()


class BatchIn(BaseModel):
    batch_id: str
    material_type: str
    raw_mass_kg: float
    concentration_percent: float
    arrival_date: date
    supplier: str | None = None
    warehouse: str | None = None
    certificate_number: str | None = None
    notes: str | None = None
    remaining_raw_mass_kg: float | None = None
    source: str = "manual"

    @field_validator("material_type")
    @classmethod
    def material_ok(cls, value: str) -> str:
        value = value.upper()
        if value not in MATERIALS:
            raise ValueError("material_type must be A, B or C")
        return value

    @field_validator("batch_id")
    @classmethod
    def id_ok(cls, value: str) -> str:
        if not value.strip():
            raise ValueError("batch_id is required")
        return value.strip()

    @field_validator("raw_mass_kg")
    @classmethod
    def mass_ok(cls, value: float) -> float:
        if value <= 0:
            raise ValueError("raw_mass_kg must be positive")
        return value

    @field_validator("concentration_percent")
    @classmethod
    def concentration_ok(cls, value: float) -> float:
        if not 0 <= value <= 100:
            raise ValueError("concentration_percent must be 0..100")
        return value


class RequirementIn(BaseModel):
    requirements: dict[str, float]
    policy: str = "hybrid"
    allow_rework: bool = True

    @field_validator("requirements")
    @classmethod
    def requirements_ok(cls, value: dict[str, float]) -> dict[str, float]:
        if any(material not in MATERIALS or amount < 0 for material, amount in value.items()):
            raise ValueError("requirements must contain only non-negative A, B or C values")
        return {material: float(amount) for material, amount in value.items()}

    @field_validator("policy")
    @classmethod
    def policy_ok(cls, value: str) -> str:
        if value not in POLICIES: raise ValueError(f"policy must be one of {POLICIES}")
        return value


class RuleIn(BaseModel):
    good_threshold_percent: float
    rework_threshold_percent: float
    good_recovery_factor: float = 1.0
    rework_recovery_factor: float = 0.9
    reject_recovery_factor: float = 0.0


class ChatIn(BaseModel):
    message: str = Field(min_length=1, max_length=4000)
    chat_id: str | None = None
    history: list[dict[str, str]] = Field(default_factory=list, max_length=20)


class AuthIn(BaseModel):
    username: str = Field(min_length=3, max_length=32)
    password: str = Field(min_length=8, max_length=128)


class UserPatch(BaseModel):
    is_blocked: bool | None = None
    is_admin: bool | None = None


class RegistrationSetting(BaseModel):
    registration_open: bool


def hash_password(password: str, salt: bytes | None = None) -> str:
    salt = salt or secrets.token_bytes(16)
    rounds = 240_000
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, rounds)
    return f"pbkdf2${rounds}${salt.hex()}${digest.hex()}"


def verify_password(password: str, encoded: str) -> bool:
    try:
        scheme, rounds, salt_hex, digest_hex = encoded.split("$", 3)
        if scheme != "pbkdf2":
            return False
        digest = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt_hex), int(rounds))
        return hmac.compare_digest(digest.hex(), digest_hex)
    except (ValueError, TypeError):
        return False


def public_user(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    return {"user_id": row["user_id"], "username": row["username"], "is_admin": bool(row["is_admin"]), "is_blocked": bool(row["is_blocked"]), "created_at": row["created_at"]}


def session_token(con: sqlite3.Connection, user_id: str) -> str:
    raw = secrets.token_urlsafe(32)
    con.execute("INSERT INTO sessions VALUES (?,?,?)", (hashlib.sha256(raw.encode()).hexdigest(), user_id, datetime.fromtimestamp(datetime.now().timestamp() + 60 * 60 * 24 * 30).isoformat()))
    return raw


def current_user(request: Request, admin: bool = False) -> sqlite3.Row:
    header = request.headers.get("authorization", "")
    token = header[7:] if header.lower().startswith("bearer ") else request.headers.get("x-session-token", "")
    if not token:
        raise HTTPException(401, "authentication required")
    con = db(); row = con.execute("SELECT u.* FROM sessions s JOIN users u ON u.user_id=s.user_id WHERE s.token_hash=? AND s.expires_at>?", (hashlib.sha256(token.encode()).hexdigest(), datetime.utcnow().isoformat())).fetchone()
    if not row:
        con.close(); raise HTTPException(401, "invalid or expired session")
    if row["is_blocked"]:
        con.close(); raise HTTPException(403, "account is blocked")
    if admin and not row["is_admin"]:
        con.close(); raise HTTPException(403, "admin access required")
    con.close()
    return row


def chat_for_user(con: sqlite3.Connection, chat_id: str, user_id: str) -> sqlite3.Row:
    row = con.execute("SELECT * FROM chats WHERE chat_id=? AND user_id=?", (chat_id, user_id)).fetchone()
    if not row:
        raise HTTPException(404, "chat not found")
    return row


def save_message(con: sqlite3.Connection, chat_id: str, role: str, content: str, tool_calls: list[dict[str, Any]] | None = None) -> None:
    now = datetime.utcnow().isoformat()
    con.execute("INSERT INTO chat_messages VALUES (?,?,?,?,?,?)", (str(uuid.uuid4()), chat_id, role, content, json.dumps(tool_calls or [], ensure_ascii=False), now))
    con.execute("UPDATE chats SET updated_at=? WHERE chat_id=?", (now, chat_id))


def rules() -> dict[str, dict[str, float]]:
    con = db(); rows = con.execute("SELECT * FROM quality_rules").fetchall(); con.close()
    return {r["material_type"]: dict(r) for r in rows}


def classify(row: dict[str, Any]) -> dict[str, Any]:
    rule = rules()[row["material_type"]]
    c = float(row["concentration_percent"])
    if c >= rule["good_threshold_percent"]:
        status, recovery = "GOOD", rule["good_recovery_factor"]
    elif c >= rule["rework_threshold_percent"]:
        status, recovery = "REWORK", rule["rework_recovery_factor"]
    else:
        status, recovery = "REJECTED", rule["reject_recovery_factor"]
    return {"status": status, "recovery_factor": recovery, "reason": {"GOOD": "Концентрация соответствует норме", "REWORK": "Концентрация ниже нормы, но партия пригодна для доработки", "REJECTED": "Концентрация ниже порога доработки"}[status]}


def batch_dict(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    result = dict(row)
    result["quality"] = classify(result)
    result["theoretical_active_mass_kg"] = round(result["raw_mass_kg"] * result["concentration_percent"] / 100, 3)
    result["available_active_mass_kg"] = round(result["remaining_raw_mass_kg"] * result["concentration_percent"] / 100 * result["quality"]["recovery_factor"], 3)
    result["meta"] = snapshot_meta({"raw_mass": "kg_raw", "active_mass": "kg_active"}, {"batch_id": result["batch_id"]})
    return result


def validate_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    valid, errors, seen = [], [], set()
    con = db(); existing = {r[0] for r in con.execute("SELECT batch_id FROM batches")}; con.close()
    for number, raw in enumerate(rows, 2):
        try:
            item = BatchIn.model_validate(raw)
            if item.batch_id in seen or item.batch_id in existing:
                raise ValueError("batch_id must be unique")
            remaining = item.raw_mass_kg if item.remaining_raw_mass_kg is None else item.remaining_raw_mass_kg
            if remaining < 0 or remaining > item.raw_mass_kg:
                raise ValueError("remaining_raw_mass_kg must be between 0 and raw_mass_kg")
            data = item.model_dump(mode="json"); data["remaining_raw_mass_kg"] = remaining
            valid.append(data); seen.add(item.batch_id)
        except Exception as exc:
            errors.append({"row": number, "field": "row", "code": "INVALID_VALUE", "message": str(exc)})
    return valid, errors


def save_batches(rows: list[dict[str, Any]]) -> int:
    con = db()
    for item in rows:
        con.execute("INSERT INTO batches VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", (
            item["batch_id"], item["material_type"], item["raw_mass_kg"], item["concentration_percent"],
            item["arrival_date"], item.get("supplier"), item.get("warehouse"), item.get("certificate_number"),
            item.get("notes"), item["remaining_raw_mass_kg"], item.get("source", "upload"), None, datetime.utcnow().isoformat()))
    bump_data_version(con); con.commit(); con.close(); return len(rows)


def parse_file(content: bytes, filename: str) -> list[dict[str, Any]]:
    if filename.lower().endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            name = next((n for n in archive.namelist() if n.lower().endswith("batches.csv")), None)
            if not name: raise ValueError("ZIP must contain batches.csv")
            content, filename = archive.read(name), name
    if filename.lower().endswith(".xlsx"):
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = book[book.sheetnames[0]]
        values = list(sheet.values); headers = [str(x) for x in values[0]]
        return [dict(zip(headers, row)) for row in values[1:] if any(x is not None for x in row)]
    text = content.decode("utf-8-sig")
    sample = text[:4096]
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    return list(csv.DictReader(io.StringIO(text), dialect=dialect))


def sort_batches(rows: list[dict[str, Any]], policy: str, allow_rework: bool) -> list[dict[str, Any]]:
    allowed = [x for x in rows if x["quality"]["status"] == "GOOD" or (allow_rework and x["quality"]["status"] == "REWORK")]
    if policy == "strict_fifo":
        return sorted(allowed, key=lambda x: (x["arrival_date"], x["created_at"], x["batch_id"]))
    if policy == "max_concentration":
        return sorted(allowed, key=lambda x: (x["quality"]["status"] != "GOOD", -x["concentration_percent"], x["arrival_date"], x["batch_id"]))
    good = sorted([x for x in allowed if x["quality"]["status"] == "GOOD"], key=lambda x: (x["arrival_date"], x["created_at"], x["batch_id"]))
    rework = sorted([x for x in allowed if x["quality"]["status"] == "REWORK"], key=lambda x: (-x["concentration_percent"], x["arrival_date"], x["batch_id"]))
    return good + rework


def build_plan(requirements: dict[str, float], policy: str = "hybrid", allow_rework: bool = True) -> dict[str, Any]:
    if policy not in POLICIES: raise ValueError(f"policy must be one of {POLICIES}")
    con = db(); all_rows = [batch_dict(r) for r in con.execute("SELECT * FROM batches WHERE remaining_raw_mass_kg > 0").fetchall()]; con.close()
    by_material = {m: [r for r in all_rows if r["material_type"] == m] for m in MATERIALS}
    result = {"policy": policy, "requirements": requirements, "materials": {}, "warnings": [], "meta": snapshot_meta({"requirements": "kg_active", "raw_mass": "kg_raw", "active_mass": "kg_active"}, {"policy": policy, "allow_rework": allow_rework, "requirements": requirements})}
    for material in MATERIALS:
        required = max(0.0, float(requirements.get(material, 0)))
        need = required; items = []; available = 0.0
        for selection_rank, row in enumerate(sort_batches(by_material[material], policy, allow_rework), 1):
            factor = row["concentration_percent"] / 100 * row["quality"]["recovery_factor"]
            capacity = row["remaining_raw_mass_kg"] * factor; available += capacity
            if need <= 1e-9: continue
            active = min(need, capacity); raw = active / factor if factor else 0
            if policy == "strict_fifo":
                selection_reason = f"Самая ранняя доступная {row['quality']['status']}-партия по FIFO"
            elif policy == "max_concentration":
                selection_reason = f"Максимальная концентрация среди доступных {row['quality']['status']}-партий"
            elif row["quality"]["status"] == "GOOD":
                selection_reason = "GOOD-партия выбрана по FIFO до использования REWORK"
            else:
                selection_reason = "REWORK-партия выбрана после GOOD по убыванию концентрации"
            items.append({"batch_id": row["batch_id"], "status": row["quality"]["status"], "selection_rank": selection_rank, "selection_reason": selection_reason, "arrival_date": row["arrival_date"], "concentration_percent": row["concentration_percent"], "raw_mass_used_kg": round(raw, 3), "active_mass_kg": round(active, 3), "active_mass_received_kg": round(active, 3), "loss_kg": round(raw * row["concentration_percent"] / 100 * (1 - row["quality"]["recovery_factor"]), 3)})
            need -= active
        covered = required - max(0, need)
        result["materials"][material] = {"required_active_mass_kg": round(required, 3), "available_active_mass_kg": round(available, 3), "covered_active_mass_kg": round(covered, 3), "deficit_active_mass_kg": round(max(0, need), 3), "coverage_percent": round(covered / required * 100, 2) if required else 100.0, "items": items, "raw_mass_used_kg": round(sum(i["raw_mass_used_kg"] for i in items), 3), "loss_kg": round(sum(i["loss_kg"] for i in items), 3), "selection_explanation": f"Стратегия {policy}; отобрано партий: {len(items)}"}
        if need > 1e-9: result["warnings"].append(f"Дефицит материала {material}: {need:.3f} кг активного вещества")
    return result


def requirements_default() -> dict[str, float]:
    con = db(); rows = con.execute("SELECT * FROM requirements").fetchall(); con.close(); return {r[0]: r[1] for r in rows}


def tool(name: str, args: dict[str, Any]) -> Any:
    con = db()
    if name == "check_batch_quality":
        row = con.execute("SELECT * FROM batches WHERE batch_id=?", (args["batch_id"],)).fetchone(); con.close()
        if not row: raise ValueError("batch not found")
        return batch_dict(row)
    if name == "get_batch_details":
        row = con.execute("SELECT * FROM batches WHERE batch_id=?", (args["batch_id"],)).fetchone(); con.close()
        if not row: raise ValueError("batch not found")
        return batch_dict(row)
    if name == "classify_batches":
        material = args.get("material_type"); rows = con.execute("SELECT * FROM batches" + (" WHERE material_type=?" if material else ""), (material,) if material else ()).fetchall(); con.close(); counts = {"GOOD": 0, "REWORK": 0, "REJECTED": 0}
        for row in rows: counts[classify(dict(row))["status"]] += 1
        return {"checked": len(rows), **counts, "run_id": str(uuid.uuid4()), "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], args)}
    if name == "get_inventory_summary":
        material = args.get("material_type"); rows = [batch_dict(r) for r in con.execute("SELECT * FROM batches" + (" WHERE material_type=?" if material else ""), (material,) if material else ()).fetchall()]; con.close(); summary = {}
        for row in rows:
            key = f'{row["material_type"]}:{row["quality"]["status"]}' if args.get("group_by", "material_and_status") == "material_and_status" else row["material_type"]
            item = summary.setdefault(key, {"material_type": row["material_type"], "status": row["quality"]["status"], "batch_count": 0, "raw_mass_kg": 0, "theoretical_active_mass_kg": 0, "available_active_mass_kg": 0, "recovery_loss_active_mass_kg": 0})
            remaining_theoretical_active = row["remaining_raw_mass_kg"] * row["concentration_percent"] / 100
            item["batch_count"] += 1; item["raw_mass_kg"] += row["remaining_raw_mass_kg"]; item["theoretical_active_mass_kg"] += remaining_theoretical_active; item["available_active_mass_kg"] += row["available_active_mass_kg"]
            item["recovery_loss_active_mass_kg"] += remaining_theoretical_active - row["available_active_mass_kg"]
        return {"groups": list(summary.values()), "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], {"material_type": material, "group_by": args.get("group_by", "material_and_status")})}
    if name == "build_weekly_plan":
        con.close(); return build_plan(args.get("requirements") or requirements_default(), args.get("policy", "hybrid"), args.get("allow_rework", True))
    if name == "check_material_deficit":
        plan = build_plan(args.get("requirements") or {}, "hybrid", args.get("include_rework", True)); return {"materials": plan["materials"], "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], args)}
    if name == "compare_allocation_policies":
        con.close(); req = args.get("requirements") or requirements_default(); return {"policies": {p: build_plan(req, p, True) for p in args.get("policies", POLICIES)}, "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], args)}
    if name == "generate_rejection_report":
        material = args.get("material_type"); rows = [batch_dict(r) for r in con.execute("SELECT * FROM batches" + (" WHERE material_type=?" if material else ""), (material,) if material else ()).fetchall()]; con.close(); statuses = set([x for x, enabled in (("REWORK", args.get("include_rework", True)), ("REJECTED", args.get("include_rejected", True))) if enabled]); selected = [r for r in rows if r["quality"]["status"] in statuses]; return {"report_id": str(uuid.uuid4()), "batches": selected, "total_batches": len(selected), "total_raw_mass_kg": sum(r["remaining_raw_mass_kg"] for r in selected), "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], args)}
    if name == "simulate_requirement_change":
        con.close(); base = args.get("base_requirements") or requirements_default(); changes = args.get("changes_percent") or {}; new = {m: base.get(m, 0) * (1 + changes.get(m, 0) / 100) for m in MATERIALS}; base_plan = build_plan(base, args.get("policy", "hybrid")); new_plan = build_plan(new, args.get("policy", "hybrid")); comparison = {}
        for material in MATERIALS:
            before = base_plan["materials"][material]; after = new_plan["materials"][material]; required = float(base.get(material, 0) or 0); safe_growth = max(0.0, (before["available_active_mass_kg"] / required - 1) * 100) if required else 0.0
            comparison[material] = {"base_deficit_active_mass_kg": before["deficit_active_mass_kg"], "new_deficit_active_mass_kg": after["deficit_active_mass_kg"], "deficit_delta_active_mass_kg": round(after["deficit_active_mass_kg"] - before["deficit_active_mass_kg"], 3), "base_rework_batches": sum(item["status"] == "REWORK" for item in before["items"]), "new_rework_batches": sum(item["status"] == "REWORK" for item in after["items"]), "safe_growth_percent": round(safe_growth, 2)}
        return {"base": base_plan, "new": new_plan, "new_requirements": new, "comparison": comparison, "meta": snapshot_meta(TOOL_REGISTRY[name]["units"], args)}
    con.close(); raise ValueError("unknown tool")


TOOLS = tuple(TOOL_REGISTRY)


def local_agent(message: str) -> tuple[str, str, Any, list[dict[str, Any]]]:
    intent = route_intent(message)
    if intent["intent"] != "EXECUTE_TOOL":
        if intent["intent"] == "CLARIFY": return "offline", "Нужно уточнить: " + ", ".join(intent.get("missing_fields", [])), None, []
        return "offline", "Я работаю с партиями, остатками, дефицитом, планом и отчётами. Для расчёта сформулируйте явную команду.", None, []
    if intent.get("tool_name") and intent.get("arguments"):
        name, args = intent["tool_name"], intent["arguments"]
        data = tool(name, args)
        return name, summarize_tool_result(data), data, [{"tool": name, "arguments": args, "status": "success"}]
    text = message.lower()
    match = re.search(r"\b([abc]-\d+)\b", text)
    old_material = next((m for m in MATERIALS if m.lower() in text and any(x in text for x in ("стар", "ранн"))), None)
    if old_material:
        con = db(); oldest = con.execute("SELECT batch_id FROM batches WHERE material_type=? ORDER BY arrival_date, created_at, batch_id LIMIT 1", (old_material,)).fetchone(); con.close()
        name, args = ("check_batch_quality", {"batch_id": oldest[0]}) if oldest else ("get_inventory_summary", {"material_type": old_material})
    elif match and any(x in text for x in ("проверь", "кач", "статус")): name, args = "check_batch_quality", {"batch_id": match.group(1).upper()}
    elif any(x in text for x in ("брак", "отбрак")): name, args = "generate_rejection_report", {}
    elif any(x in text for x in ("сравни", "fifo", "стратег")): name, args = "compare_allocation_policies", {"requirements": requirements_default()}
    elif any(x in text for x in ("план", "составь")): name, args = "build_weekly_plan", {"requirements": requirements_default()}
    elif any(x in text for x in ("хватит", "дефицит", "потребн")): name, args = "check_material_deficit", {"requirements": requirements_default(), "include_rework": True}
    else: name, args = "get_inventory_summary", {}
    data = tool(name, args)
    if name == "check_batch_quality":
        answer = f"Партия {data['batch_id']}: {data['quality']['status']}, концентрация {data['concentration_percent']}%. {data['quality']['reason']}."
    else:
        answer = f"Режим без LLM: выполнен инструмент {name}. Результат доступен в блоке расчёта."
    return name, answer, data, [{"tool": name, "arguments": args, "status": "success"}]


def tool_specs() -> list[dict[str, Any]]:
    return [{"type": "function", "function": {"name": name, "description": spec["description"], "parameters": {"type": "object", "properties": spec["parameters"], "required": spec["required"], "additionalProperties": False}}} for name, spec in TOOL_REGISTRY.items()]


def summarize_tool_result(data: Any) -> str:
    if isinstance(data, dict) and data.get("groups") is not None:
        groups = data["groups"]
        if not groups:
            return "Остатков по текущему фильтру не найдено."
        lines = ["Сводка по остаткам:"]
        for item in groups:
            lines.append(f"{item['material_type']} / {item['status']}: {item['batch_count']} партий, {item['raw_mass_kg']:.1f} кг сырья, {item['available_active_mass_kg']:.1f} кг активного вещества, потери восстановления {item.get('recovery_loss_active_mass_kg', 0):.1f} кг")
        return "\n".join(lines)
    if isinstance(data, dict) and data.get("batch_id"):
        quality = data.get("quality", {})
        return f"Партия {data['batch_id']}: {quality.get('status', 'без статуса')}, концентрация {data.get('concentration_percent', 0)}%. {quality.get('reason', '')}".strip()
    if isinstance(data, dict) and data.get("materials"):
        lines = ["Preview-план сформирован:"]
        for material, item in data["materials"].items():
            lines.append(f"{material}: покрытие {item['coverage_percent']:.1f}%, дефицит {item['deficit_active_mass_kg']:.1f} кг активного вещества")
        return "\n".join(lines)
    if isinstance(data, dict) and data.get("batches") is not None:
        return f"Отчёт сформирован: {len(data['batches'])} проблемных партий, {data.get('total_raw_mass_kg', 0):.1f} кг остатка к решению."
    if isinstance(data, dict) and data.get("comparison"):
        lines = ["Сценарий рассчитан:"]
        for material, item in data["comparison"].items():
            lines.append(f"{material}: дефицит {item['base_deficit_active_mass_kg']:.1f} → {item['new_deficit_active_mass_kg']:.1f} кг активного вещества; безопасный рост {item['safe_growth_percent']:.1f}%")
        return "\n".join(lines)
    return "Расчёт завершён. Подробности доступны в результате инструмента."


def result_numbers(data: Any) -> set[float]:
    numbers: set[float] = set()
    if isinstance(data, dict):
        for value in data.values(): numbers.update(result_numbers(value))
    elif isinstance(data, (int, float)) and not isinstance(data, bool):
        numbers.add(round(float(data), 6))
    elif isinstance(data, list):
        for value in data: numbers.update(result_numbers(value))
    return numbers


def answer_numbers_are_grounded(answer: str, data: Any) -> bool:
    if not data: return True
    expected = result_numbers(data)
    for token in re.findall(r"(?<![A-Za-zА-Яа-яЁё0-9-])\d+(?:[.,]\d+)?(?![A-Za-zА-Яа-яЁё0-9-])", answer or ""):
        value = float(token.replace(",", "."))
        if not any(abs(value - candidate) <= max(0.051, abs(candidate) * 0.00001) for candidate in expected): return False
    return True


def registry_explanation(tool_name: str) -> str:
    spec = TOOL_REGISTRY[tool_name]
    params = ", ".join(spec["parameters"]) or "без обязательных параметров"
    return f"{spec['title']} ({tool_name}) — {spec['description']} Параметры: {params}. Инструмент read-only: производственные данные не изменяет."


def forced_tool_for_message(message: str, history: list[dict[str, str]] | None = None) -> tuple[str, dict[str, Any]] | None:
    intent = route_intent(message, history)
    if intent["intent"] != "EXECUTE_TOOL": return None
    if intent.get("tool_name") and intent.get("arguments"):
        return intent["tool_name"], intent["arguments"]
    normalized = message.lower()
    material = next((item for item in MATERIALS if re.search(rf"\b{item.lower()}\b", normalized)), None)
    if any(word in normalized for word in ("остат", "склад", "сырь")):
        return "get_inventory_summary", {"material_type": material, "group_by": "material_and_status"}
    if any(word in normalized for word in ("брак", "отбрак", "доработ")):
        return "generate_rejection_report", {"material_type": material, "include_rework": True, "include_rejected": True}
    batch_match = re.search(r"\b([abc]-\d+)\b", normalized)
    if batch_match and any(word in normalized for word in ("проверь", "качество", "статус", "парт")):
        return "check_batch_quality", {"batch_id": batch_match.group(1).upper()}
    if any(word in normalized for word in ("недельный план", "составь план", "план производства", "построй план")):
        requirements = parse_requirements(message)
        if set(requirements) == set(MATERIALS): return "build_weekly_plan", {"requirements": requirements, "policy": "hybrid", "allow_rework": True}
    return None


def requested_material(message: str) -> str | None:
    normalized = message.lower()
    if re.search(r"\b(ашки|а-шки)\b", normalized): return "A"
    if re.search(r"\bб\b", normalized): return "B"
    if re.search(r"\bс\b", normalized): return "C"
    return next((item for item in MATERIALS if re.search(rf"\b{item.lower()}\b", normalized)), None)


def material_choices(prompt: str) -> list[dict[str, str]]:
    return [{"label": "Все материалы", "value": f"{prompt} по всем материалам"}] + [{"label": item, "value": f"{prompt} по материалу {item}"} for item in MATERIALS]


def parse_requirement_details(message: str) -> tuple[dict[str, float], str | None]:
    normalized = message.lower().replace(",", ".")
    values: dict[str, float] = {}
    units = {"т": 1000.0, "тонн": 1000.0, "тонны": 1000.0, "kg": 1.0, "кг": 1.0, "г": 0.001, "g": 0.001}
    found_unit: str | None = None
    for material in MATERIALS:
        match = re.search(rf"(?:материал\s*)?\b{material.lower()}\b(?:\s+(?:на\s+)?(?:потребность|нужно|нужна|спрос)\s*)?\s*(?:[:=]|[-–])?\s*(\d+(?:\.\d+)?)\s*(т|тонн|тонны|кг|kg|г|g)?", normalized)
        if match:
            unit = match.group(2) or "кг"
            values[material] = float(match.group(1)) * units[unit]
            found_unit = unit
    return values, found_unit


def parse_requirements(message: str) -> dict[str, float]:
    return parse_requirement_details(message)[0]


def parse_changes(message: str) -> dict[str, float]:
    normalized = message.lower().replace(",", ".")
    changes: dict[str, float] = {}
    for material in MATERIALS:
        match = re.search(rf"\b{material.lower()}\b[^%]{{0,45}}?(?:на|рост\w*|выраст\w*|увелич\w*)\s*([+-]?\d+(?:\.\d+)?)\s*%", normalized)
        if match: changes[material] = float(match.group(1))
    return changes


def mass_basis(message: str) -> str | None:
    normalized = message.lower()
    if any(word in normalized for word in ("активн", "действующ", "active")):
        return "active_mass_kg"
    if any(word in normalized for word in ("сыр", "raw")):
        return "raw_mass_kg"
    return None


def clarification_text(intent: dict[str, Any]) -> str:
    missing = intent.get("missing_fields", [])
    if "mass_basis" in missing:
        return "Уточните, числа указаны как масса активного вещества или как масса сырья."
    return "Уточните, пожалуйста: " + ", ".join(missing or ["параметры запроса"]) + "."


def route_intent(message: str, history: list[dict[str, str]] | None = None) -> dict[str, Any]:
    normalized = message.lower()
    explanation = explanation_tool_for_message(message)
    if explanation:
        return {"intent": "EXPLAIN_TOOL", "tool_name": explanation, "arguments": {}, "missing_fields": [], "confidence": 0.99, "reason": "Запрос содержит просьбу объяснить инструмент"}
    if any(word in normalized for word in ("открой обзор", "открой партии", "открой план", "открой отчёт", "открой отчет", "открой ассистент")):
        target = "data" if "парт" in normalized else "planning" if "план" in normalized else "reports" if "отч" in normalized else "assistant" if "ассист" in normalized else "overview"
        return {"intent": "NAVIGATE", "tool_name": None, "arguments": {"page": target}, "missing_fields": [], "confidence": 0.98, "reason": "Явная команда навигации"}
    context = "\n".join(item.get("content", "") for item in (history or [])[-8:]).lower()
    previous_user = next((item.get("content", "") for item in reversed(history or []) if item.get("role") == "user"), "")
    previous_requirements = parse_requirements(previous_user)
    if history and requested_material(message) and any(word in normalized for word in ("а теперь", "теперь", "ещё", "еще")) and any(word in context for word in ("остат", "склад", "запас", "inventory")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "get_inventory_summary", "arguments": {"material_type": requested_material(message), "group_by": "material_and_status"}, "missing_fields": [], "confidence": 0.91, "reason": "Материал уточнён по контексту предыдущего запроса"}
    if history and previous_requirements and set(previous_requirements) == set(MATERIALS) and any(word in normalized for word in ("такой же", "как до этого", "как раньше", "но fifo", "но hybrid", "тоже")) and "план" in context:
        policy = "strict_fifo" if "fifo" in normalized else "max_concentration" if "концентрац" in normalized else "hybrid"
        return {"intent": "EXECUTE_TOOL", "tool_name": "build_weekly_plan", "arguments": {"requirements": previous_requirements, "policy": policy, "allow_rework": True, "mass_basis": mass_basis(previous_user) or "active_mass_kg"}, "missing_fields": [], "confidence": 0.9, "reason": "Параметры плана унаследованы из контекста"}
    if history and previous_requirements and "план" in context and any(word in normalized for word in ("%", "процент", "больше", "увеличь", "рост")):
        changes = {material.upper(): float(value) for material, value in re.findall(r"\b([abc])\b\s*(?:на\s*)?([+-]?\d+(?:[.,]\d+)?)\s*%", normalized)}
        policy = "strict_fifo" if "fifo" in context else "max_concentration" if "концентрац" in context else "hybrid"
        if changes:
            return {"intent": "EXECUTE_TOOL", "tool_name": "simulate_requirement_change", "arguments": {"base_requirements": previous_requirements, "changes_percent": changes, "policy": policy}, "missing_fields": [], "confidence": 0.88, "reason": "Сценарий изменения унаследовал базовую потребность из контекста"}
    material = requested_material(message)
    batch_match = re.search(r"\b([abc]-\d+)\b", normalized)
    if batch_match and any(word in normalized for word in ("детал", "покажи", "карточ")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "get_batch_details", "arguments": {"batch_id": batch_match.group(1).upper()}, "missing_fields": [], "confidence": 0.99, "reason": "Явно запрошены детали партии"}
    if batch_match and any(word in normalized for word in ("проверь", "качество", "статус", "парт")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "check_batch_quality", "arguments": {"batch_id": batch_match.group(1).upper()}, "missing_fields": [], "confidence": 0.99, "reason": "Явно указана партия и команда проверки"}
    if any(word in normalized for word in ("классифиц", "классификац")) and any(word in normalized for word in ("запусти", "сделай", "проведи", "покажи", "выполни")):
        return {"intent": "EXECUTE_TOOL", "tool_name": "classify_batches", "arguments": {"material_type": material, "only_unclassified": False}, "missing_fields": [], "confidence": 0.96, "reason": "Явно запрошена классификация партий"}
    if any(word in normalized for word in ("остат", "остал", "склад", "запас", "сырь", "че по", "потер")) and any(word in normalized for word in ("покажи", "покаж", "дай", "проверь", "посчитай", "сколько", "че по", "где")):
        if not material and not any(word in normalized for word in ("все", "всем", "общ", "где")):
            return {"intent": "CLARIFY", "tool_name": "get_inventory_summary", "arguments": {}, "missing_fields": ["material_type"], "confidence": 0.95, "reason": "Нужно выбрать материал или все материалы", "choices": material_choices("Покажи остатки")}
        return {"intent": "EXECUTE_TOOL", "tool_name": "get_inventory_summary", "arguments": {"material_type": material, "group_by": "material_and_status"}, "missing_fields": [], "confidence": 0.98, "reason": "Явно запрошена сводка остатков"}
    if any(word in normalized for word in ("брак", "отбрак", "доработ", "проблемн", "отклон")) and any(word in normalized for word in ("покажи", "сделай", "сформируй", "отчёт", "отчет", "парти")):
        if not material and not any(word in normalized for word in ("все", "всем")):
            return {"intent": "CLARIFY", "tool_name": "generate_rejection_report", "arguments": {}, "missing_fields": ["material_type"], "confidence": 0.95, "reason": "Нужно выбрать материал или все материалы", "choices": material_choices("Покажи отчёт по браку")}
        return {"intent": "EXECUTE_TOOL", "tool_name": "generate_rejection_report", "arguments": {"material_type": material, "include_rework": True, "include_rejected": True}, "missing_fields": [], "confidence": 0.98, "reason": "Явно запрошен отчёт по браку"}
    if any(word in normalized for word in ("недельный план", "составь план", "построй план", "план производства")):
        requirements = parse_requirements(message); missing = [m for m in MATERIALS if m not in requirements]
        policy = "strict_fifo" if "fifo" in normalized and "hybrid" not in normalized else "max_concentration" if any(alias in normalized for alias in ("концентрац", "max concentration", "max_concentration")) else "hybrid" if "hybrid" in normalized else None
        missing_fields = [f"requirements.{m}" for m in missing] + ([] if policy else ["policy"])
        if not missing and not mass_basis(message): missing_fields.append("mass_basis")
        if missing_fields:
            return {"intent": "CLARIFY", "tool_name": "build_weekly_plan", "arguments": {"requirements": requirements}, "missing_fields": missing_fields, "confidence": 0.94, "reason": "Для preview-плана не хватает обязательных вводных", "choices": [{"label": "Hybrid", "value": "Построй недельный план с policy hybrid"}, {"label": "Strict FIFO", "value": "Построй недельный план с policy strict_fifo"}, {"label": "Max concentration", "value": "Построй недельный план с policy max_concentration"}]}
        return {"intent": "EXECUTE_TOOL", "tool_name": "build_weekly_plan", "arguments": {"requirements": requirements, "policy": policy, "allow_rework": True, "mass_basis": mass_basis(message)}, "missing_fields": [], "confidence": 0.98, "reason": "Все обязательные параметры preview-плана указаны"}
    if any(word in normalized for word in ("дефицит", "хватит", "потребн", "вытян", "достат")) and any(word in normalized for word in ("проверь", "посчитай", "есть", "хватит", "вытян", "достат")):
        requirements = parse_requirements(message)
        if not requirements: return {"intent": "CLARIFY", "tool_name": "check_material_deficit", "arguments": {}, "missing_fields": ["requirements"], "confidence": 0.9, "reason": "Нужна потребность в кг активного вещества"}
        if not mass_basis(message): return {"intent": "CLARIFY", "tool_name": "check_material_deficit", "arguments": {"requirements": requirements}, "missing_fields": ["mass_basis"], "confidence": 0.94, "reason": "Нужно различить массу сырья и активного вещества"}
        return {"intent": "EXECUTE_TOOL", "tool_name": "check_material_deficit", "arguments": {"requirements": requirements, "include_rework": True, "mass_basis": mass_basis(message)}, "missing_fields": [], "confidence": 0.96, "reason": "Указана потребность для проверки дефицита"}
    if any(word in normalized for word in ("сравни", "стратег", "политик")) and any(word in normalized for word in ("fifo", "hybrid", "концентрац", "стратег")):
        requirements = parse_requirements(message)
        if not requirements: return {"intent": "CLARIFY", "tool_name": "compare_allocation_policies", "arguments": {}, "missing_fields": ["requirements"], "confidence": 0.9, "reason": "Нужна потребность для честного сравнения стратегий"}
        if not mass_basis(message): return {"intent": "CLARIFY", "tool_name": "compare_allocation_policies", "arguments": {"requirements": requirements}, "missing_fields": ["mass_basis"], "confidence": 0.94, "reason": "Нужно различить массу сырья и активного вещества"}
        policies = [p for p, aliases in (("strict_fifo", ("fifo", "strict_fifo")), ("hybrid", ("hybrid",)), ("max_concentration", ("концентрац", "max_concentration"))) if any(alias in normalized for alias in aliases)] or list(POLICIES)
        return {"intent": "EXECUTE_TOOL", "tool_name": "compare_allocation_policies", "arguments": {"requirements": requirements, "policies": policies, "mass_basis": mass_basis(message)}, "missing_fields": [], "confidence": 0.96, "reason": "Указаны потребность и стратегии"}
    if any(word in normalized for word in ("сценар", "смоделируй", "изменение спроса", "что будет, если", "рост")):
        changes = parse_changes(message)
        policy = "strict_fifo" if "fifo" in normalized and "hybrid" not in normalized else "max_concentration" if any(alias in normalized for alias in ("концентрац", "max concentration", "max_concentration")) else "hybrid" if "hybrid" in normalized else None
        missing_fields = ([] if changes else ["changes_percent"]) + ([] if policy else ["policy"])
        if missing_fields:
            return {"intent": "CLARIFY", "tool_name": "simulate_requirement_change", "arguments": {"changes_percent": changes}, "missing_fields": missing_fields, "confidence": 0.92, "reason": "Для сценария не хватает процента изменения и политики", "choices": [{"label": "Hybrid", "value": "Смоделируй сценарий с policy hybrid"}, {"label": "Strict FIFO", "value": "Смоделируй сценарий с policy strict_fifo"}]}
        return {"intent": "EXECUTE_TOOL", "tool_name": "simulate_requirement_change", "arguments": {"base_requirements": {}, "changes_percent": changes, "policy": policy}, "missing_fields": [], "confidence": 0.95, "reason": "Указаны параметры сценария"}
    if any(word in normalized for word in ("привет", "что ты умеешь", "помоги", "что можешь", "как пользоваться")):
        return {"intent": "GENERAL_HELP", "tool_name": None, "arguments": {}, "missing_fields": [], "confidence": 0.98, "reason": "Общий вопрос без бизнес-команды"}
    return {"intent": "GENERAL_HELP", "tool_name": None, "arguments": {}, "missing_fields": [], "confidence": 0.55, "reason": "Не найдено однозначной команды на выполнение"}


def explanation_tool_for_message(message: str) -> str | None:
    normalized = message.lower()
    if not any(word in normalized for word in ("как работает", "как устроен", "объясни", "расскажи про", "что делает", "зачем нужен", "какие параметры", "как рассчитывается", "не запускай", "только объясни")):
        return None
    for name, spec in TOOL_REGISTRY.items():
        aliases = (spec["title"].lower(), name.replace("_", " "), *spec.get("aliases", []))
        if any(alias in normalized for alias in aliases): return name
    return None


def compact_history(history: list[dict[str, str]]) -> list[dict[str, str]]:
    """Keep only the small conversational window needed for an ambiguous request."""
    compact = []
    for item in history[-6:]:
        content = (item.get("content") or "").strip()
        if content:
            compact.append({"role": item["role"], "content": content[:700]})
    return compact


def llm_context(history: list[dict[str, str]]) -> str:
    last_user = next((item["content"] for item in reversed(history) if item.get("role") == "user"), "")
    return json.dumps({
        "last_user_request": last_user[:700],
        "last_requirements": parse_requirements(last_user),
        "last_material": requested_material(last_user),
        "last_policy": "strict_fifo" if "fifo" in last_user.lower() else "max_concentration" if "концентрац" in last_user.lower() else "hybrid" if "hybrid" in last_user.lower() else None,
    }, ensure_ascii=False)


def routed_tool_result(message: str, history: list[dict[str, str]]) -> tuple[str, Any, list[dict[str, Any]]] | None:
    intent = route_intent(message, history)
    if intent["intent"] != "EXECUTE_TOOL" or not intent.get("tool_name") or not intent.get("arguments"):
        return None
    name, args = intent["tool_name"], intent["arguments"]
    try:
        data = tool(name, args)
        trace = [{"tool": name, "arguments": args, "status": "success", "source": "router"}]
    except Exception as exc:
        data = {"error": str(exc)}
        trace = [{"tool": name, "arguments": args, "status": "error", "message": str(exc), "source": "router"}]
    return name, data, trace


def explanation_prompt(message: str, history: list[dict[str, str]], tool_name: str | None = None, data: Any = None) -> tuple[str, list[dict[str, str]]]:
    system = """Ты — AI-технолог системы контроля качества сырья. Отвечай по-русски, коротко и по делу. Не придумывай числа, партии или статусы. Если ниже есть результат инструмента, объясни именно его простыми словами и отделяй массу сырья от массы активного вещества. Не запускай инструменты и не меняй данные."""
    if tool_name and data is not None:
        prompt = f"Запрос пользователя: {message}\nИнструмент {tool_name} уже выполнен. Его результат JSON:\n{json.dumps(data, ensure_ascii=False, separators=(',', ':'))[:24000]}"
    else:
        prompt = message
    messages = [{"role": "system", "content": system + "\nСжатый контекст: " + llm_context(history)}]
    messages.extend(compact_history(history))
    messages.append({"role": "user", "content": prompt})
    return system, messages


def llm_agent(message: str, history: list[dict[str, str]]) -> tuple[str, str, Any, list[dict[str, Any]]] | None:
    key = os.getenv("LLM_API_KEY")
    if not key: return None
    base = os.getenv("LLM_BASE_URL", "https://api.openai.com/v1").rstrip("/")
    intent = route_intent(message, history)
    if intent["intent"] == "CLARIFY": return "assistant", clarification_text(intent), {"choices": intent.get("choices", [])}, []
    routed = routed_tool_result(message, history)
    if routed or intent["intent"] in ("EXPLAIN_TOOL", "GENERAL_HELP"):
        name, data, trace = routed or (None, None, [])
        _, messages = explanation_prompt(message, history, name, data)
        body = {"model": os.getenv("LLM_MODEL", "openai/gpt-5-nano"), "temperature": float(os.getenv("LLM_TEMPERATURE", "0.1")), "max_tokens": int(os.getenv("LLM_EXPLAIN_MAX_TOKENS", "650")), "messages": messages}
        req = URLRequest(base + "/chat/completions", data=json.dumps(body).encode(), headers={"Authorization": "Bearer " + key, "Content-Type": "application/json"}, method="POST")
        with urlopen(req, timeout=float(os.getenv("LLM_TIMEOUT_SECONDS", "30"))) as response: result = json.load(response)
        content = (result["choices"][0]["message"].get("content") or "").strip()
        if data and not answer_numbers_are_grounded(content, data): content = summarize_tool_result(data)
        return "llm", content or summarize_tool_result(data), data, trace
    explanation_tool = explanation_tool_for_message(message)
    system = """Ты — AI-технолог системы контроля качества сырья. Отвечай по-русски. Все данные и расчёты получай только через инструменты. Не придумывай партии и числа. После tool call объясни результат простыми словами, отделяй массу сырья от массы активного вещества. Preview-план не подтверждай сам. Всегда заверши ответ коротким понятным текстом; пустой ответ запрещён."""
    if explanation_tool: system += f" Пользователь просит объяснить инструмент {explanation_tool}. {registry_explanation(explanation_tool)} Расскажи назначение, когда он нужен, параметры и результат. Не запускай инструмент и не выдумывай поля. Ответ краткий — до 120 слов."
    messages = [{"role": "system", "content": system + "\nСжатый контекст: " + llm_context(history)}]
    messages.extend(compact_history(history))
    messages.append({"role": "user", "content": message})
    trace: list[dict[str, Any]] = []
    last_data: Any = None
    forced_used = False
    retry_used = False
    for _ in range(int(os.getenv("LLM_MAX_TOOL_ROUNDS", "4"))):
        body = {"model": os.getenv("LLM_MODEL", "openai/gpt-5-nano"), "temperature": float(os.getenv("LLM_TEMPERATURE", "0.1")), "max_tokens": int(os.getenv("LLM_MAX_TOKENS", "2000")), "messages": messages, "tools": tool_specs(), "tool_choice": "auto"}
        req = URLRequest(base + "/chat/completions", data=json.dumps(body).encode(), headers={"Authorization": "Bearer " + key, "Content-Type": "application/json"}, method="POST")
        with urlopen(req, timeout=float(os.getenv("LLM_TIMEOUT_SECONDS", "30"))) as response: result = json.load(response)
        msg = result["choices"][0]["message"]
        messages.append(msg)
        calls = msg.get("tool_calls") or []
        if calls and intent["intent"] != "EXECUTE_TOOL":
            content = (msg.get("content") or "").strip() or (registry_explanation(explanation_tool) if explanation_tool else "Я могу показать фактические данные через инструменты, но для этого нужна явная команда на расчёт.")
            return "llm", content, None, trace
        if not calls:
            forced = forced_tool_for_message(message, history) if not forced_used else None
            if forced:
                forced_used = True
                name, args = forced; call_id = "forced-" + uuid.uuid4().hex
                try:
                    last_data = tool(name, args); trace.append({"tool": name, "arguments": args, "status": "success", "source": "forced"}); tool_result = last_data
                except Exception as exc:
                    trace.append({"tool": name, "arguments": args, "status": "error", "message": str(exc), "source": "forced"}); tool_result = {"error": str(exc)}
                messages.append({"role": "assistant", "content": None, "tool_calls": [{"id": call_id, "type": "function", "function": {"name": name, "arguments": json.dumps(args)}}]})
                messages.append({"role": "tool", "tool_call_id": call_id, "name": name, "content": json.dumps(tool_result, ensure_ascii=False)})
                continue
            content = (msg.get("content") or "").strip()
            if not content and not retry_used:
                retry_used = True
                messages.append({"role": "user", "content": "Сформулируй короткий содержательный ответ на исходный вопрос. Не возвращай пустой ответ."})
                continue
            if last_data and not answer_numbers_are_grounded(content, last_data): content = summarize_tool_result(last_data)
            return "llm", content or summarize_tool_result(last_data), last_data, trace
        for call in calls:
            name = call["function"]["name"]
            if name not in TOOLS: raise ValueError("unknown tool")
            args = json.loads(call["function"].get("arguments") or "{}")
            try:
                last_data = tool(name, args); trace.append({"tool": name, "arguments": args, "status": "success", "source": "model"})
                tool_result = last_data
            except Exception as exc:
                trace.append({"tool": name, "arguments": args, "status": "error", "message": str(exc), "source": "model"}); tool_result = {"error": str(exc)}
            messages.append({"role": "tool", "tool_call_id": call["id"], "name": name, "content": json.dumps(tool_result, ensure_ascii=False)})
    return "llm", "Не удалось завершить расчёт за допустимое число шагов.", last_data, trace


def llm_agent_stream(message: str, history: list[dict[str, str]]):
    intent = route_intent(message, history)
    if intent["intent"] == "CLARIFY":
        answer = clarification_text(intent)
        yield {"type": "token", "text": answer}
        yield {"type": "done", "response": {"mode": "assistant", "answer": answer, "result": {"choices": intent.get("choices", [])}, "choices": intent.get("choices", []), "needs_clarification": True, "tool_calls": []}}
        return
    key = os.getenv("LLM_API_KEY")
    if not key:
        name, answer, data, trace = local_agent(message)
        yield {"type": "done", "response": {"mode": "offline", "answer": answer, "result": data, "tool_calls": trace}}
        return
    base = os.getenv("LLM_BASE_URL", "https://api.openai.com/v1").rstrip("/")
    routed = routed_tool_result(message, history)
    if routed or intent["intent"] in ("EXPLAIN_TOOL", "GENERAL_HELP"):
        name, data, trace = routed or (None, None, [])
        if trace:
            yield {"type": "tool", "tool": trace[-1]["tool"], "status": trace[-1]["status"]}
        _, messages = explanation_prompt(message, history, name, data)
        body = {"model": os.getenv("LLM_MODEL", "openai/gpt-5-nano"), "temperature": float(os.getenv("LLM_TEMPERATURE", "0.1")), "max_tokens": int(os.getenv("LLM_EXPLAIN_MAX_TOKENS", "650")), "messages": messages, "stream": True}
        req = URLRequest(base + "/chat/completions", data=json.dumps(body).encode(), headers={"Authorization": "Bearer " + key, "Content-Type": "application/json"}, method="POST")
        answer_parts: list[str] = []
        with urlopen(req, timeout=float(os.getenv("LLM_TIMEOUT_SECONDS", "30"))) as response:
            for raw_line in response:
                line = raw_line.decode("utf-8", errors="ignore").strip()
                if not line.startswith("data:"): continue
                payload = line[5:].strip()
                if payload == "[DONE]": break
                delta = json.loads(payload).get("choices", [{}])[0].get("delta", {}).get("content") or ""
                if delta: answer_parts.append(delta)
        answer = "".join(answer_parts).strip()
        if data and not answer_numbers_are_grounded(answer, data): answer = summarize_tool_result(data)
        answer = answer or summarize_tool_result(data)
        yield {"type": "token", "text": answer}
        yield {"type": "done", "response": {"mode": "llm", "answer": answer, "result": data, "tool_calls": trace}}
        return
    explanation_tool = explanation_tool_for_message(message)
    system = """Ты — AI-технолог системы контроля качества сырья. Отвечай по-русски. Все данные и расчёты получай только через инструменты. Не придумывай партии и числа. После tool call объясни результат простыми словами, отделяй массу сырья от массы активного вещества. Preview-план не подтверждай сам. Всегда заверши ответ коротким понятным текстом; пустой ответ запрещён."""
    if explanation_tool: system += f" Пользователь просит объяснить инструмент {explanation_tool}. {registry_explanation(explanation_tool)} Расскажи назначение, когда он нужен, параметры и результат. Не запускай инструмент и не выдумывай поля. Ответ краткий — до 120 слов."
    messages = [{"role": "system", "content": system + "\nСжатый контекст: " + llm_context(history)}]
    messages.extend(compact_history(history))
    messages.append({"role": "user", "content": message})
    trace: list[dict[str, Any]] = []
    last_data: Any = None
    forced_used = False
    retry_used = False
    max_rounds = int(os.getenv("LLM_MAX_TOOL_ROUNDS", "2"))
    for round_index in range(max_rounds):
        body = {"model": os.getenv("LLM_MODEL", "openai/gpt-5-nano"), "temperature": float(os.getenv("LLM_TEMPERATURE", "0.1")), "max_tokens": int(os.getenv("LLM_MAX_TOKENS", "2000")), "messages": messages, "tools": tool_specs(), "tool_choice": "auto"}
        streaming = round_index > 0
        if streaming:
            body["stream"] = True
            body["tool_choice"] = "auto" if explanation_tool else "none"
        req = URLRequest(base + "/chat/completions", data=json.dumps(body).encode(), headers={"Authorization": "Bearer " + key, "Content-Type": "application/json"}, method="POST")
        with urlopen(req, timeout=float(os.getenv("LLM_TIMEOUT_SECONDS", "30"))) as response:
            if not streaming:
                result = json.load(response)
                msg = result["choices"][0]["message"]
                messages.append(msg)
                calls = msg.get("tool_calls") or []
                if calls and intent["intent"] != "EXECUTE_TOOL":
                    answer = (msg.get("content") or "").strip() or (registry_explanation(explanation_tool) if explanation_tool else "Я могу показать фактические данные через инструменты, но для этого нужна явная команда на расчёт.")
                    yield {"type": "token", "text": answer}
                    yield {"type": "done", "response": {"mode": "llm", "answer": answer, "result": None, "tool_calls": []}}
                    return
                if not calls:
                    forced = forced_tool_for_message(message, history) if not forced_used else None
                    if forced:
                        forced_used = True
                        name, args = forced; call_id = "forced-" + uuid.uuid4().hex
                        try:
                            last_data = tool(name, args); trace.append({"tool": name, "arguments": args, "status": "success", "source": "forced"}); tool_result = last_data
                        except Exception as exc:
                            trace.append({"tool": name, "arguments": args, "status": "error", "message": str(exc), "source": "forced"}); tool_result = {"error": str(exc)}
                        messages.append({"role": "assistant", "content": None, "tool_calls": [{"id": call_id, "type": "function", "function": {"name": name, "arguments": json.dumps(args)}}]})
                        messages.append({"role": "tool", "tool_call_id": call_id, "name": name, "content": json.dumps(tool_result, ensure_ascii=False)})
                        yield {"type": "tool", "tool": name, "status": trace[-1]["status"]}
                        continue
                    content = (msg.get("content") or "").strip()
                    if not content and not retry_used:
                        retry_used = True
                        messages.append({"role": "user", "content": "Сформулируй короткий содержательный ответ на исходный вопрос. Не возвращай пустой ответ."})
                        continue
                    if last_data and not answer_numbers_are_grounded(content, last_data): content = summarize_tool_result(last_data)
                    answer = content or summarize_tool_result(last_data)
                    yield {"type": "token", "text": answer}
                    yield {"type": "done", "response": {"mode": "llm", "answer": answer, "result": last_data, "tool_calls": trace}}
                    return
                for call in calls:
                    name = call["function"]["name"]
                    if name not in TOOLS: raise ValueError("unknown tool")
                    args = json.loads(call["function"].get("arguments") or "{}")
                    try:
                        last_data = tool(name, args); trace.append({"tool": name, "arguments": args, "status": "success", "source": "model"}); tool_result = last_data
                    except Exception as exc:
                        trace.append({"tool": name, "arguments": args, "status": "error", "message": str(exc), "source": "model"}); tool_result = {"error": str(exc)}
                    yield {"type": "tool", "tool": name, "status": trace[-1]["status"]}
                    messages.append({"role": "tool", "tool_call_id": call["id"], "name": name, "content": json.dumps(tool_result, ensure_ascii=False)})
                continue
            answer_parts: list[str] = []
            for raw_line in response:
                line = raw_line.decode("utf-8", errors="ignore").strip()
                if not line.startswith("data:"): continue
                payload = line[5:].strip()
                if payload == "[DONE]": break
                chunk = json.loads(payload)
                delta = chunk.get("choices", [{}])[0].get("delta", {})
                text = delta.get("content") or ""
                if text:
                    answer_parts.append(text)
            answer = "".join(answer_parts).strip()
            if last_data and not answer_numbers_are_grounded(answer, last_data): answer = summarize_tool_result(last_data)
            answer = answer or summarize_tool_result(last_data)
            yield {"type": "token", "text": answer}
            yield {"type": "done", "response": {"mode": "llm", "answer": answer, "result": last_data, "tool_calls": trace}}
            return
    answer = "Не удалось завершить расчёт за допустимое число шагов."
    yield {"type": "token", "text": answer}
    yield {"type": "done", "response": {"mode": "llm", "answer": answer, "result": last_data, "tool_calls": trace}}


def sse_event(payload: dict[str, Any]) -> str:
    return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"


app = FastAPI(title="Raw Material AI", version="0.1.0")
app.add_middleware(CORSMiddleware, allow_origins=os.getenv("CORS_ORIGINS", "*").split(","), allow_methods=["*"], allow_headers=["*"])


@app.on_event("startup")
def startup() -> None: init_db()


@app.get("/", include_in_schema=False)
def index() -> FileResponse: return FileResponse(Path(__file__).parent / "web" / "index.html")


@app.post("/api/v1/auth/register")
def register(payload: AuthIn) -> dict[str, Any]:
    username = payload.username.strip()
    if not re.fullmatch(r"[A-Za-zА-Яа-яЁё0-9_.-]{3,32}", username):
        raise HTTPException(422, "username contains unsupported characters")
    con = db()
    if con.execute("SELECT value FROM settings WHERE key='registration_open'").fetchone()[0] != "1":
        con.close(); raise HTTPException(403, "registration is closed")
    if con.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
        con.close(); raise HTTPException(409, "username already exists")
    user_id = str(uuid.uuid4())
    con.execute("INSERT INTO users VALUES (?,?,?,?,?,?)", (user_id, username, hash_password(payload.password), 0, 0, datetime.utcnow().isoformat()))
    token = session_token(con, user_id); con.commit(); row = con.execute("SELECT * FROM users WHERE user_id=?", (user_id,)).fetchone(); con.close()
    return {"token": token, "user": public_user(row)}


@app.post("/api/v1/auth/login")
def login(payload: AuthIn) -> dict[str, Any]:
    con = db(); row = con.execute("SELECT * FROM users WHERE username=?", (payload.username.strip(),)).fetchone()
    if not row or not verify_password(payload.password, row["password_hash"]):
        con.close(); raise HTTPException(401, "invalid username or password")
    if row["is_blocked"]:
        con.close(); raise HTTPException(403, "account is blocked")
    token = session_token(con, row["user_id"]); con.commit(); con.close()
    return {"token": token, "user": public_user(row)}


@app.post("/api/v1/auth/logout")
def logout(request: Request) -> dict[str, bool]:
    header = request.headers.get("authorization", ""); token = header[7:] if header.lower().startswith("bearer ") else request.headers.get("x-session-token", "")
    con = db(); con.execute("DELETE FROM sessions WHERE token_hash=?", (hashlib.sha256(token.encode()).hexdigest(),)); con.commit(); con.close(); return {"ok": True}


@app.get("/api/v1/auth/me")
def me(request: Request) -> dict[str, Any]: return {"user": public_user(current_user(request))}


@app.get("/api/v1/chats")
def chats(request: Request) -> list[dict[str, Any]]:
    user = current_user(request); con = db(); rows = con.execute("SELECT chat_id, title, created_at, updated_at FROM chats WHERE user_id=? ORDER BY updated_at DESC", (user["user_id"],)).fetchall(); con.close(); return [dict(r) for r in rows]


@app.post("/api/v1/chats")
def create_chat(request: Request, title: str = "Новый чат") -> dict[str, Any]:
    user = current_user(request); now = datetime.utcnow().isoformat(); item = {"chat_id": str(uuid.uuid4()), "title": title[:80] or "Новый чат", "created_at": now, "updated_at": now}; con = db(); con.execute("INSERT INTO chats VALUES (?,?,?,?,?)", (item["chat_id"], user["user_id"], item["title"], now, now)); con.commit(); con.close(); return item


@app.get("/api/v1/chats/{chat_id}/messages")
def chat_messages(request: Request, chat_id: str) -> list[dict[str, Any]]:
    user = current_user(request); con = db(); chat_for_user(con, chat_id, user["user_id"]); rows = con.execute("SELECT role, content, tool_calls, created_at FROM chat_messages WHERE chat_id=? ORDER BY created_at", (chat_id,)).fetchall(); con.close(); return [{**dict(r), "tool_calls": json.loads(r["tool_calls"] or "[]")} for r in rows]


@app.get("/api/v1/admin/settings")
def admin_settings(request: Request) -> dict[str, bool]:
    current_user(request, admin=True); con = db(); value = con.execute("SELECT value FROM settings WHERE key='registration_open'").fetchone()[0] == "1"; con.close(); return {"registration_open": value}


@app.put("/api/v1/admin/settings")
def admin_settings_update(payload: RegistrationSetting, request: Request) -> dict[str, bool]:
    current_user(request, admin=True); con = db(); con.execute("UPDATE settings SET value=? WHERE key='registration_open'", ("1" if payload.registration_open else "0",)); con.commit(); con.close(); return {"registration_open": payload.registration_open}


@app.get("/api/v1/admin/users")
def admin_users(request: Request) -> list[dict[str, Any]]:
    current_user(request, admin=True); con = db(); rows = con.execute("SELECT user_id, username, is_admin, is_blocked, created_at FROM users ORDER BY created_at").fetchall(); con.close(); return [public_user(r) for r in rows]


@app.patch("/api/v1/admin/users/{user_id}")
def admin_user_update(user_id: str, payload: UserPatch, request: Request) -> dict[str, Any]:
    admin = current_user(request, admin=True); con = db(); row = con.execute("SELECT * FROM users WHERE user_id=?", (user_id,)).fetchone()
    if not row: con.close(); raise HTTPException(404, "user not found")
    if user_id == admin["user_id"] and payload.is_blocked:
        con.close(); raise HTTPException(400, "cannot block current admin")
    if payload.is_blocked is not None: con.execute("UPDATE users SET is_blocked=? WHERE user_id=?", (int(payload.is_blocked), user_id))
    if payload.is_admin is not None: con.execute("UPDATE users SET is_admin=? WHERE user_id=?", (int(payload.is_admin), user_id))
    if payload.is_blocked: con.execute("DELETE FROM sessions WHERE user_id=?", (user_id,))
    con.commit(); row = con.execute("SELECT * FROM users WHERE user_id=?", (user_id,)).fetchone(); con.close(); return public_user(row)


@app.get("/api/v1/admin/chats")
def admin_chats(request: Request) -> list[dict[str, Any]]:
    current_user(request, admin=True); con = db(); rows = con.execute("SELECT c.chat_id, c.title, c.created_at, c.updated_at, u.username FROM chats c JOIN users u ON u.user_id=c.user_id ORDER BY c.updated_at DESC").fetchall(); con.close(); return [dict(r) for r in rows]


@app.get("/api/v1/admin/chats/{chat_id}/messages")
def admin_chat_messages(chat_id: str, request: Request) -> list[dict[str, Any]]:
    current_user(request, admin=True); con = db(); rows = con.execute("SELECT role, content, tool_calls, created_at FROM chat_messages WHERE chat_id=? ORDER BY created_at", (chat_id,)).fetchall(); con.close(); return [{**dict(r), "tool_calls": json.loads(r["tool_calls"] or "[]")} for r in rows]


@app.get("/health")
def health() -> dict[str, str]: return {"status": "ok"}


@app.get("/ready")
def ready() -> dict[str, str]:
    con = db(); con.execute("SELECT 1"); con.close(); return {"status": "ready"}


@app.get("/api/v1/agent/tools")
def agent_tools(request: Request) -> list[dict[str, Any]]:
    current_user(request)
    return [{"name": name, **spec} for name, spec in TOOL_REGISTRY.items()]


@app.get("/api/v1/batches")
def batches(request: Request, material_type: str | None = None) -> list[dict[str, Any]]:
    current_user(request)
    con = db(); rows = con.execute("SELECT * FROM batches" + (" WHERE material_type=?" if material_type else "") + " ORDER BY arrival_date, batch_id", (material_type,) if material_type else ()).fetchall(); con.close(); return [batch_dict(r) for r in rows]


@app.post("/api/v1/batches")
def add_batch(item: BatchIn, request: Request) -> dict[str, Any]:
    current_user(request)
    valid, errors = validate_rows([item.model_dump(mode="json")]);
    if errors: raise HTTPException(422, errors)
    save_batches(valid); con = db(); row = con.execute("SELECT * FROM batches WHERE batch_id=?", (item.batch_id,)).fetchone(); con.close(); return batch_dict(row)


@app.get("/api/v1/batches/{batch_id}")
def get_batch(batch_id: str, request: Request) -> dict[str, Any]:
    current_user(request)
    return tool("get_batch_details", {"batch_id": batch_id})


@app.delete("/api/v1/batches/{batch_id}")
def delete_batch(batch_id: str, request: Request) -> dict[str, bool]:
    current_user(request)
    con = db(); cur = con.execute("DELETE FROM batches WHERE batch_id=?", (batch_id,));
    if cur.rowcount: bump_data_version(con)
    con.commit(); con.close();
    if not cur.rowcount: raise HTTPException(404, "batch not found")
    return {"deleted": True}


@app.get("/api/v1/quality-rules")
def get_rules(request: Request) -> list[dict[str, Any]]:
    current_user(request)
    return list(rules().values())


@app.put("/api/v1/quality-rules/{material_type}")
def put_rule(material_type: str, item: RuleIn, request: Request) -> dict[str, Any]:
    current_user(request)
    if material_type not in MATERIALS or item.rework_threshold_percent > item.good_threshold_percent: raise HTTPException(422, "invalid rule")
    con = db(); con.execute("UPDATE quality_rules SET good_threshold_percent=?, rework_threshold_percent=?, good_recovery_factor=?, rework_recovery_factor=?, reject_recovery_factor=? WHERE material_type=?", (*item.model_dump().values(), material_type)); bump_data_version(con); con.commit(); con.close(); return {"material_type": material_type, **item.model_dump()}


@app.post("/api/v1/classifications/run")
def classify_api(request: Request, material_type: str | None = None) -> Any:
    current_user(request)
    return tool("classify_batches", {"material_type": material_type})


@app.get("/api/v1/inventory/summary")
def inventory(request: Request, material_type: str | None = None) -> Any:
    current_user(request)
    return tool("get_inventory_summary", {"material_type": material_type})


@app.post("/api/v1/imports/preview")
async def import_preview(request: Request, file: UploadFile = File(...)) -> Any:
    current_user(request)
    try: rows = parse_file(await file.read(), file.filename or "upload.csv"); valid, errors = validate_rows(rows); return {"valid_rows": len(valid), "invalid_rows": len(errors), "errors": errors, "rows": valid}
    except Exception as exc: raise HTTPException(400, str(exc))


class ImportCommit(BaseModel): rows: list[dict[str, Any]]


@app.post("/api/v1/imports/commit")
def import_commit(payload: ImportCommit, request: Request) -> dict[str, int]:
    current_user(request)
    valid, errors = validate_rows(payload.rows)
    if errors: raise HTTPException(422, {"errors": errors})
    return {"imported": save_batches(valid)}


@app.post("/api/v1/plans/preview")
def plan_preview(payload: RequirementIn, request: Request) -> Any:
    user = current_user(request); plan_id = str(uuid.uuid4()); plan = build_plan(payload.requirements, payload.policy, payload.allow_rework)
    con = db(); con.execute("INSERT INTO plan_owners(plan_id,user_id,status,created_at,preview_data_version) VALUES (?,?,?,?,?)", (plan_id, user["user_id"], "preview", datetime.utcnow().isoformat(), plan["meta"]["data_version"])); con.commit(); con.close()
    return {"plan_id": plan_id, **plan}


@app.post("/api/v1/plans/{plan_id}/confirm")
def plan_confirm(plan_id: str, payload: RequirementIn, request: Request) -> Any:
    user = current_user(request)
    con = db(); owner = con.execute("SELECT status, preview_data_version FROM plan_owners WHERE plan_id=? AND user_id=?", (plan_id, user["user_id"])).fetchone()
    if not owner: con.close(); raise HTTPException(404, "plan not found")
    if owner["status"] != "preview": con.close(); raise HTTPException(409, "plan already confirmed or cancelled")
    current_version_row = con.execute("SELECT value FROM settings WHERE key='data_version'").fetchone()
    current_version = int(current_version_row[0]) if current_version_row else 1
    if int(owner["preview_data_version"]) != current_version:
        con.close(); raise HTTPException(409, {"code": "STALE_PLAN", "message": "Складские остатки изменились. Пересчитайте план.", "preview_data_version": int(owner["preview_data_version"]), "current_data_version": current_version})
    existing = con.execute("SELECT status FROM plans WHERE plan_id=?", (plan_id,)).fetchone()
    if existing: con.close(); raise HTTPException(409, "plan already confirmed or cancelled")
    plan = build_plan(payload.requirements, payload.policy, payload.allow_rework)
    try:
        con.execute("INSERT INTO plans VALUES (?,?,?,?)", (plan_id, "confirmed", json.dumps(plan), datetime.utcnow().isoformat()))
        for item in [i for m in plan["materials"].values() for i in m["items"]]:
            cur = con.execute("UPDATE batches SET remaining_raw_mass_kg=remaining_raw_mass_kg-? WHERE batch_id=? AND remaining_raw_mass_kg>=?", (item["raw_mass_used_kg"], item["batch_id"], item["raw_mass_used_kg"]))
            if cur.rowcount != 1: raise HTTPException(409, {"code": "STALE_PLAN", "message": "Складские остатки изменились. Пересчитайте план."})
        con.execute("UPDATE plan_owners SET status='confirmed' WHERE plan_id=? AND user_id=? AND status='preview'", (plan_id, user["user_id"]))
        bump_data_version(con); con.commit()
    except HTTPException:
        con.rollback(); con.close(); raise
    except Exception:
        con.rollback(); con.close(); raise
    con.close(); return {"plan_id": plan_id, "status": "confirmed", "plan": plan}


@app.post("/api/v1/policies/compare")
def compare(payload: RequirementIn, request: Request) -> Any:
    current_user(request)
    return tool("compare_allocation_policies", {"requirements": payload.requirements, "policies": POLICIES})


@app.post("/api/v1/scenarios/simulate")
def simulate(payload: dict[str, Any], request: Request) -> Any:
    current_user(request)
    return tool("simulate_requirement_change", payload)


@app.post("/api/v1/agent/chat")
def chat(payload: ChatIn, request: Request) -> Any:
    user = current_user(request)
    con = db()
    if payload.chat_id:
        chat_row = chat_for_user(con, payload.chat_id, user["user_id"])
        chat_id = chat_row["chat_id"]
    else:
        chat_id = str(uuid.uuid4()); now = datetime.utcnow().isoformat(); title = payload.message.strip()[:80]
        con.execute("INSERT INTO chats VALUES (?,?,?,?,?)", (chat_id, user["user_id"], title or "Новый чат", now, now))
    previous = con.execute("SELECT role, content FROM chat_messages WHERE chat_id=? ORDER BY created_at DESC LIMIT 12", (chat_id,)).fetchall()
    history = [{"role": r["role"], "content": r["content"]} for r in reversed(previous)] or payload.history
    save_message(con, chat_id, "user", payload.message)
    con.execute("UPDATE chats SET title=? WHERE chat_id=? AND title='Новый чат'", (payload.message.strip()[:80] or "Новый чат", chat_id)); con.commit(); con.close()
    try:
        normalized = payload.message.lower(); explanation = explanation_tool_for_message(payload.message)
        intent = route_intent(payload.message, history)
        if intent["intent"] == "CLARIFY":
            answer = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": clarification_text(intent), "question": clarification_text(intent), "needs_clarification": True, "choices": intent.get("choices", []), "tool_calls": []}
            con = db(); save_message(con, chat_id, answer["answer"]); con.commit(); con.close(); return answer
        if any(x in normalized for x in ("проверь партию", "карточк* партии")) and not re.search(r"\b[abc]-\d+\b", normalized) and "сам" not in normalized and not explanation:
            con = db(); choices = [r[0] for r in con.execute("SELECT batch_id FROM batches ORDER BY arrival_date, batch_id LIMIT 20")]; con.close()
            answer = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "Какую партию проверить?", "question": "Какую партию проверить?", "needs_clarification": True, "choices": choices, "tool_calls": []}
            con = db(); save_message(con, chat_id, "assistant", answer["answer"]); con.commit(); con.close(); return answer
        if "самую стар" in normalized and not any(m.lower() in normalized for m in MATERIALS) and not explanation:
            answer = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "Для какого материала найти самую старую партию?", "question": "Для какого материала найти самую старую партию?", "needs_clarification": True, "choices": list(MATERIALS), "tool_calls": []}
            con = db(); save_message(con, chat_id, "assistant", answer["answer"]); con.commit(); con.close(); return answer
        if any(x in normalized for x in ("отчет по браку", "отчёт по браку", "покажи брак")) and not any(m.lower() in normalized for m in MATERIALS) and "все" not in normalized and not explanation:
            answer = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "По какому материалу сформировать отчёт?", "question": "По какому материалу сформировать отчёт?", "needs_clarification": True, "choices": material_choices("Покажи отчёт по браку"), "tool_calls": []}
            con = db(); save_message(con, chat_id, "assistant", answer["answer"]); con.commit(); con.close(); return answer
        material = requested_material(payload.message)
        plan_request = any(word in normalized for word in ("недельный план", "составь план", "план производства", "построй план")) and not explanation
        if plan_request:
            supplied = parse_requirements(payload.message); missing = [m for m in MATERIALS if m not in supplied]
            if missing:
                answer = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": f"Чтобы построить недельный план, укажите потребность по активному веществу для {', '.join(missing)}. Например: A 3000, B 3000, C 3000.", "needs_clarification": True, "choices": [], "tool_calls": []}
                con = db(); save_message(con, chat_id, "assistant", answer["answer"]); con.commit(); con.close(); return answer
        if any(word in normalized for word in ("остат", "склад", "запас")) and not material and "все" not in normalized and "всем" not in normalized and not explanation:
            answer = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "По какому материалу показать остатки?", "question": "По какому материалу показать остатки?", "needs_clarification": True, "choices": material_choices("Покажи остатки"), "tool_calls": []}
            con = db(); save_message(con, chat_id, "assistant", answer["answer"]); con.commit(); con.close(); return answer
        if any(word in normalized for word in ("дефицит", "хватит", "потребн")) and not material and "все" not in normalized and "всем" not in normalized and not explanation:
            answer = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "Для какого материала проверить дефицит?", "question": "Для какого материала проверить дефицит?", "needs_clarification": True, "choices": material_choices("Проверь дефицит"), "tool_calls": []}
            con = db(); save_message(con, chat_id, "assistant", answer["answer"]); con.commit(); con.close(); return answer
        if any(word in normalized for word in ("сравни стратег", "сравнить стратег")) and not any(word in normalized for word in ("fifo", "hybrid", "концентрац")) and not explanation:
            answer = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "Какие стратегии сравнить?", "question": "Какие стратегии сравнить?", "needs_clarification": True, "choices": [{"label": "Все три стратегии", "value": "Сравни все стратегии распределения"}, {"label": "FIFO vs Hybrid", "value": "Сравни стратегии FIFO и hybrid"}, {"label": "FIFO vs концентрация", "value": "Сравни strict_fifo и max_concentration"}], "tool_calls": []}
            con = db(); save_message(con, chat_id, "assistant", answer["answer"]); con.commit(); con.close(); return answer
        result = llm_agent(payload.message, history) or local_agent(payload.message)
        name, answer, data, trace = result
        response = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "llm" if name == "llm" else "offline", "answer": answer, "tool": trace[-1]["tool"] if trace else None, "tool_calls": trace, "result": data}
        con = db(); save_message(con, chat_id, "assistant", answer, trace); con.commit(); con.close(); return response
    except Exception as exc:
        con = db(); save_message(con, chat_id, "assistant", f"Не удалось выполнить запрос: {exc}"); con.commit(); con.close(); raise HTTPException(502, str(exc))


@app.post("/api/v1/agent/chat/stream")
def chat_stream(payload: ChatIn, request: Request) -> StreamingResponse:
    user = current_user(request)
    con = db()
    if payload.chat_id:
        chat_id = chat_for_user(con, payload.chat_id, user["user_id"])["chat_id"]
    else:
        chat_id = str(uuid.uuid4()); now = datetime.utcnow().isoformat(); title = payload.message.strip()[:80]
        con.execute("INSERT INTO chats VALUES (?,?,?,?,?)", (chat_id, user["user_id"], title or "Новый чат", now, now))
    previous = con.execute("SELECT role, content FROM chat_messages WHERE chat_id=? ORDER BY created_at DESC LIMIT 12", (chat_id,)).fetchall()
    history = [{"role": r["role"], "content": r["content"]} for r in reversed(previous)] or payload.history
    save_message(con, chat_id, "user", payload.message)
    con.execute("UPDATE chats SET title=? WHERE chat_id=? AND title='Новый чат'", (payload.message.strip()[:80] or "Новый чат", chat_id)); con.commit(); con.close()

    def events():
        response: dict[str, Any] | None = None
        try:
            normalized = payload.message.lower(); explanation = explanation_tool_for_message(payload.message)
            intent = route_intent(payload.message, history)
            if intent["intent"] == "CLARIFY":
                answer = clarification_text(intent)
                response = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": answer, "question": answer, "needs_clarification": True, "choices": intent.get("choices", []), "tool_calls": []}
                yield sse_event({"type": "token", "text": answer})
            elif any(x in normalized for x in ("проверь партию", "карточк* партии")) and not re.search(r"\b[abc]-\d+\b", normalized) and "сам" not in normalized and not explanation:
                con = db(); choices = [r[0] for r in con.execute("SELECT batch_id FROM batches ORDER BY arrival_date, batch_id LIMIT 20")]; con.close()
                response = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "Какую партию проверить?", "question": "Какую партию проверить?", "needs_clarification": True, "choices": choices, "tool_calls": []}
                yield sse_event({"type": "token", "text": response["answer"]})
            elif "самую стар" in normalized and not any(m.lower() in normalized for m in MATERIALS) and not explanation:
                response = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "Для какого материала найти самую старую партию?", "question": "Для какого материала найти самую старую партию?", "needs_clarification": True, "choices": list(MATERIALS), "tool_calls": []}
                yield sse_event({"type": "token", "text": response["answer"]})
            elif any(x in normalized for x in ("отчет по браку", "отчёт по браку", "покажи брак")) and not any(m.lower() in normalized for m in MATERIALS) and "все" not in normalized and not explanation:
                response = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "По какому материалу сформировать отчёт?", "question": "По какому материалу сформировать отчёт?", "needs_clarification": True, "choices": material_choices("Покажи отчёт по браку"), "tool_calls": []}
                yield sse_event({"type": "token", "text": response["answer"]})
            elif any(word in normalized for word in ("недельный план", "составь план", "план производства", "построй план")) and not explanation and set(parse_requirements(payload.message)) != set(MATERIALS):
                missing = [m for m in MATERIALS if m not in parse_requirements(payload.message)]
                response = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": f"Чтобы построить недельный план, укажите потребность по активному веществу для {', '.join(missing)}. Например: A 3000, B 3000, C 3000.", "needs_clarification": True, "choices": [], "tool_calls": []}
                yield sse_event({"type": "token", "text": response["answer"]})
            elif any(word in normalized for word in ("остат", "склад", "запас")) and not requested_material(payload.message) and "все" not in normalized and "всем" not in normalized and not explanation:
                response = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "По какому материалу показать остатки?", "question": "По какому материалу показать остатки?", "needs_clarification": True, "choices": material_choices("Покажи остатки"), "tool_calls": []}
                yield sse_event({"type": "token", "text": response["answer"]})
            elif any(word in normalized for word in ("дефицит", "хватит", "потребн")) and not requested_material(payload.message) and "все" not in normalized and "всем" not in normalized and not explanation:
                response = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "Для какого материала проверить дефицит?", "question": "Для какого материала проверить дефицит?", "needs_clarification": True, "choices": material_choices("Проверь дефицит"), "tool_calls": []}
                yield sse_event({"type": "token", "text": response["answer"]})
            elif any(word in normalized for word in ("сравни стратег", "сравнить стратег")) and not any(word in normalized for word in ("fifo", "hybrid", "концентрац")) and not explanation:
                response = {"run_id": str(uuid.uuid4()), "chat_id": chat_id, "mode": "assistant", "answer": "Какие стратегии сравнить?", "question": "Какие стратегии сравнить?", "needs_clarification": True, "choices": [{"label": "Все три стратегии", "value": "Сравни все стратегии распределения"}, {"label": "FIFO vs Hybrid", "value": "Сравни стратегии FIFO и hybrid"}, {"label": "FIFO vs концентрация", "value": "Сравни strict_fifo и max_concentration"}], "tool_calls": []}
                yield sse_event({"type": "token", "text": response["answer"]})
            else:
                yield sse_event({"type": "status", "text": "Запрашиваю расчёт…"})
                for item in llm_agent_stream(payload.message, history):
                    if item.get("type") == "done":
                        response = item["response"]; response.update({"run_id": str(uuid.uuid4()), "chat_id": chat_id})
                    else:
                        yield sse_event(item)
            if response:
                con = db(); save_message(con, chat_id, "assistant", response["answer"], response.get("tool_calls", [])); con.commit(); con.close()
                yield sse_event({"type": "done", "response": response})
        except Exception as exc:
            try:
                con = db(); save_message(con, chat_id, "assistant", f"Не удалось выполнить запрос: {exc}"); con.commit(); con.close()
            except Exception:
                pass
            yield sse_event({"type": "error", "message": str(exc)})

    return StreamingResponse(events(), media_type="text/event-stream", headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"})


@app.get("/api/v1/reports/rejections")
def rejection_report(request: Request, material_type: str | None = None) -> Any:
    current_user(request)
    return tool("generate_rejection_report", {"material_type": material_type})


@app.get("/api/v1/reports/{report_id}/download")
def report_download(report_id: str, request: Request, material_type: str | None = None) -> StreamingResponse:
    current_user(request)
    report = tool("generate_rejection_report", {"material_type": material_type}); rows = report["batches"]; output = io.StringIO(); writer = csv.DictWriter(output, fieldnames=["batch_id", "material_type", "concentration_percent", "status", "remaining_raw_mass_kg"]); writer.writeheader(); writer.writerows({k: r.get(k) if k != "status" else r["quality"]["status"] for k in writer.fieldnames} for r in rows); return StreamingResponse(iter([output.getvalue().encode("utf-8-sig")]), media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="rejection-{report_id}.csv"'})


@app.get("/api/v1/requirements")
def get_requirements(request: Request) -> dict[str, float]:
    current_user(request)
    return requirements_default()


@app.put("/api/v1/requirements")
def put_requirements(payload: dict[str, float], request: Request) -> dict[str, float]:
    current_user(request)
    con = db();
    for material, value in payload.items():
        if material not in MATERIALS or value < 0: raise HTTPException(422, "invalid requirement")
        con.execute("INSERT INTO requirements VALUES (?,?) ON CONFLICT(material_type) DO UPDATE SET required_active_mass_kg=excluded.required_active_mass_kg", (material, value))
    bump_data_version(con); con.commit(); con.close(); return requirements_default()
